import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, timedelta
import db
import os
from PIL import Image
import numpy as np
import json
import unicodedata

# --- CSS opcional ---
st.markdown("""
<style>
    .logo-container {
        display: flex;
        justify-content: center;
        align-items: center;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# ===================== FUNCIONES DE LOGOS =====================
def get_cliente_logo_path(nombre_cliente, custom_logo_filename=None):
    # Normalizar nombre: eliminar espacios duplicados y convertir a mayúsculas
    nombre_norm = ' '.join(nombre_cliente.strip().upper().split())
    if "ONGC" in nombre_norm:
        nombre_norm = "ONGC VIDESH"
        
    mapping = {
        "APC": "logo_alk.png",
        "ARROW EXPLORATION": "logo_arrow.png",
        "CARRAO": "logo_carrao.png",
        "ECOPETROL": "logo_ecopetrol.png",
        "ECP - GPA PPU": "logo_ecopetrol.png",
        "ECP - GOR": "logo_ecopetrol.png",
        "ECP - VRC": "logo_ecopetrol.png",
        "ECP - VRO": "logo_ecopetrol.png",
        "FRONTERA": "logo_frontera.png",
        "GEOPARK": "logo_geopark.png",
        "GRAN TIERRA": "logo_grantierra.png",
        "IHSA": "logo_ihsa.png",
        "ONGC Videsh": "logo_ongc.png",
        "PAREX": "logo_parex.png",
        "SAN AGUSTIN ENERGY": "logo_sanagustin.png",
        "SIERRACOL": "logo_sierracol.png"
    }
        
    # Buscar en mapping con normalización de claves
    for key, file in mapping.items():
        key_norm = ' '.join(key.upper().split())
        if key_norm == nombre_norm or key == nombre_cliente:
            logo_path = os.path.join(os.path.dirname(__file__), "..", "assets", file)
            if os.path.exists(logo_path):
                return logo_path
            break  # Si no existe el archivo, seguir a personalizado
    
    # Si no está en mapping o el archivo por defecto no existe, buscar logo personalizado
    if custom_logo_filename and isinstance(custom_logo_filename, str) and custom_logo_filename.strip():
        custom_path = os.path.join(os.path.dirname(__file__), "..", "assets", "logos_clientes", custom_logo_filename)
        if os.path.exists(custom_path):
            return custom_path
    return None

def cargar_logo_alta_calidad(nombre_cliente, ancho_deseado=80, custom_logo_filename=None):
    logo_path = get_cliente_logo_path(nombre_cliente, custom_logo_filename)
    if not logo_path:
        return None
    try:
        img = Image.open(logo_path)
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            img = img.convert('RGBA')
        else:
            img = img.convert('RGB')
        ratio = ancho_deseado / float(img.size[0])
        alto_deseado = int(float(img.size[1]) * ratio)
        img_redim = img.resize((ancho_deseado, alto_deseado), Image.Resampling.LANCZOS)
        return img_redim
    except Exception as e:
        st.warning(f"Error cargando logo: {e}")
        return None

def mostrar_logo_cliente(nombre_cliente, width=80, centered=False, custom_logo=None):
    img = cargar_logo_alta_calidad(nombre_cliente, width, custom_logo)
    if img:
        if centered:
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.image(img, width='content')
        else:
            st.image(img, width='content')
    else:
        st.markdown("📛")

def mostrar_logo_grande(nombre_cliente, size=150, custom_logo=None):
    img = cargar_logo_alta_calidad(nombre_cliente, size, custom_logo)
    if img:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image(img, width='content')
    else:
        st.markdown("<div style='text-align: center; font-size: 4rem;'>📛</div>", unsafe_allow_html=True)

# ===================== VISTA PRINCIPAL =====================
def mostrar():
    st.title("📋 Panel de Coordinador HSE")
    st.markdown("Gestión de empleados, cursos, asignaciones e indicadores")

    tabs = st.tabs(["👥 Empleados", "🏢 Clientes y Cursos", "📚 Asignar Capacitaciones", "📊 Indicadores", "📈 Informe Gerencial", "📅 Proyección del Cronograma", "📊 Cumplimiento por Periodo DAX"])

    # --------------------- EMPLEADOS ---------------------
    with tabs[0]:
        st.subheader("Gestión de Empleados")
        col1, col2 = st.columns([1, 2])
        with col1:
            with st.expander("➕ Agregar nuevo empleado", expanded=False):
                with st.form(key="form_agregar_empleado"):
                    nueva_cedula = st.text_input("Cédula")
                    nuevo_nombre = st.text_input("Nombre completo")
                    
                    # Obtener roles desde la base de datos
                    roles_disponibles = db.obtener_roles()
                    nuevo_rol = st.selectbox("Rol", roles_disponibles, key="nuevo_rol_select")
                    
                    submitted = st.form_submit_button("Guardar empleado")
                    if submitted:
                        cedula = nueva_cedula.strip()
                        nombre = nuevo_nombre.strip()
                        if cedula and nombre:
                            if db.agregar_empleado(cedula, nombre, nuevo_rol):
                                st.success("Empleado agregado correctamente")
                                st.rerun()
                            else:
                                st.error("Cédula ya existe")
                        else:
                            st.warning("Complete todos los campos")
                
                # Agregar nuevo rol (fuera del form para no interferir con el guardado)
                with st.expander("➕ Agregar nuevo rol", expanded=False):
                    nuevo_rol_nombre = st.text_input("Nombre del nuevo rol", key="nuevo_rol_input")
                    if st.button("Crear rol", key="crear_rol_btn"):
                        if nuevo_rol_nombre.strip():
                            if db.agregar_rol(nuevo_rol_nombre.strip()):
                                st.success(f"Rol '{nuevo_rol_nombre}' creado correctamente")
                                st.rerun()
                            else:
                                st.error("El rol ya existe")
                        else:
                            st.warning("Ingresa un nombre válido")
        with col2:
            empleados_df = db.obtener_empleados()
            if not empleados_df.empty:
                st.dataframe(empleados_df[['cedula', 'nombre', 'rol']], width='stretch', hide_index=True)
                with st.expander("✏️ Editar o eliminar empleado"):
                    empleados_df = db.obtener_empleados()
                    if not empleados_df.empty:
                        emp_seleccionado = st.selectbox(
                            "Seleccionar empleado",
                            empleados_df['id'].tolist(),
                            format_func=lambda x: f"{empleados_df[empleados_df['id']==x]['nombre'].iloc[0]} ({empleados_df[empleados_df['id']==x]['cedula'].iloc[0]})",
                            key="emp_seleccionado",
                            index=None
                        )
                        if emp_seleccionado:
                            emp_data = empleados_df[empleados_df['id'] == emp_seleccionado].iloc[0]
                            nuevo_nombre = st.text_input("Nombre", value=emp_data['nombre'], key=f"edit_emp_nombre_{emp_seleccionado}")
                            nueva_cedula = st.text_input("Cédula", value=emp_data['cedula'], key=f"edit_emp_cedula_{emp_seleccionado}")
                            
                            # Obtener roles desde la base de datos
                            roles_disponibles = db.obtener_roles()
                            # Asegurar que el rol actual esté en la lista (por si se eliminó o cambió)
                            if emp_data['rol'] not in roles_disponibles:
                                roles_disponibles.append(emp_data['rol'])
                                roles_disponibles = sorted(roles_disponibles)
                            
                            nuevo_rol = st.selectbox(
                                "Rol",
                                roles_disponibles,
                                index=roles_disponibles.index(emp_data['rol']) if emp_data['rol'] in roles_disponibles else 0,
                                key=f"edit_emp_rol_{emp_seleccionado}"
                            )
                            
                            col_a, col_b = st.columns(2)
                            with col_a:
                                if st.button("Actualizar", key="actualizar_empleado"):
                                    if db.actualizar_empleado(emp_seleccionado, nueva_cedula, nuevo_nombre, nuevo_rol):
                                        st.success("Actualizado")
                                        st.rerun()
                                    else:
                                        st.error("Error: cédula duplicada")
                            with col_b:
                                if st.button("Eliminar", type="primary", key="eliminar_empleado"):
                                    db.eliminar_empleado(emp_seleccionado)
                                    st.success("Eliminado")
                                    st.rerun()
                    else:
                        st.caption("No hay empleados registrados")

    # --------------------- CLIENTES Y CURSOS ---------------------
    with tabs[1]:
        st.subheader("Clientes y Cursos por Cliente")
        col_cli, col_cursos = st.columns([1, 2])

        # Columna izquierda: Clientes
        with col_cli:
            st.markdown("### Clientes")
            clientes_df = db.obtener_clientes()
            if not clientes_df.empty:
                for idx, row in clientes_df.iterrows():
                    col_logo, col_nombre = st.columns([1, 3])
                    with col_logo:
                        logo_filename = row['logo_filename'] if pd.notna(row['logo_filename']) else None
                        mostrar_logo_cliente(row['nombre'], width=50, custom_logo=logo_filename)
                    with col_nombre:
                        st.markdown(f"**{row['nombre']}**  \n*{row['codigo']}*")
                st.divider()
            with st.expander("➕ Nuevo cliente"):
                nuevo_nom = st.text_input("Nombre cliente", key="nuevo_cliente_nom")
                nuevo_cod = st.text_input("Código", key="nuevo_cliente_cod")
                if st.button("Crear cliente", key="crear_cliente"):
                    if nuevo_nom and nuevo_cod:
                        if db.agregar_cliente(nuevo_nom, nuevo_cod):
                            st.success("Cliente agregado")
                            st.rerun()
                        else:
                            st.error("Código duplicado")
            with st.expander("✏️ Editar/Eliminar cliente"):
                if not clientes_df.empty:
                    cli_seleccionado = st.selectbox("Cliente", clientes_df['id'].tolist(),
                                                    format_func=lambda x: f"{clientes_df[clientes_df['id']==x]['nombre'].iloc[0]} ({clientes_df[clientes_df['id']==x]['codigo'].iloc[0]})",
                                                    key="cli_seleccionado")
                    if cli_seleccionado:
                        cli_row = clientes_df[clientes_df['id'] == cli_seleccionado].iloc[0]
                        st.markdown("**Logo actual:**")
                        logo_actual = cli_row['logo_filename'] if pd.notna(cli_row['logo_filename']) else None
                        mostrar_logo_grande(cli_row['nombre'], size=100, custom_logo=logo_actual)
                        
                        # Subir nuevo logo (con contador para evitar bucle)
                        if 'upload_counter' not in st.session_state:
                            st.session_state.upload_counter = 0
                        uploaded_file = st.file_uploader(
                            "Subir nuevo logo (PNG o JPG)",
                            type=['png', 'jpg', 'jpeg'],
                            key=f"upload_logo_{cli_seleccionado}_{st.session_state.upload_counter}"
                        )
                        if uploaded_file is not None:
                            logos_dir = os.path.join(os.path.dirname(__file__), "..", "assets", "logos_clientes")
                            os.makedirs(logos_dir, exist_ok=True)
                            ext = uploaded_file.name.split('.')[-1]
                            logo_filename = f"cliente_{cli_seleccionado}.{ext}"
                            logo_path = os.path.join(logos_dir, logo_filename)
                            with open(logo_path, "wb") as f:
                                f.write(uploaded_file.getbuffer())
                            db.actualizar_logo_cliente(cli_seleccionado, logo_filename)
                            st.success("Logo actualizado correctamente")
                            st.session_state.upload_counter += 1
                            st.rerun()
                        
                        ed_nombre = st.text_input("Nombre", value=cli_row['nombre'], key="edit_cli_nombre")
                        ed_codigo = st.text_input("Código", value=cli_row['codigo'], key="edit_cli_codigo")
                        col_up, col_del = st.columns(2)
                        with col_up:
                            if st.button("Actualizar cliente", key="actualizar_cliente"):
                                if db.actualizar_cliente(cli_seleccionado, ed_nombre, ed_codigo):
                                    st.success("Actualizado")
                                    st.rerun()
                                else:
                                    st.error("Código duplicado")
                        with col_del:
                            if st.button("Eliminar cliente", type="primary", key="eliminar_cliente"):
                                if cli_row['logo_filename'] and pd.notna(cli_row['logo_filename']):
                                    logo_path = os.path.join(os.path.dirname(__file__), "..", "assets", "logos_clientes", cli_row['logo_filename'])
                                    if os.path.exists(logo_path):
                                        os.remove(logo_path)
                                db.eliminar_cliente(cli_seleccionado)
                                st.success("Cliente eliminado")
                                st.rerun()
                else:
                    st.caption("No hay clientes")

        # Columna derecha: Cursos
        with col_cursos:
            st.markdown("### Cursos por Cliente")
            if not clientes_df.empty:
                cliente_para_cursos = st.selectbox("Filtrar por cliente", options=[0] + clientes_df['id'].tolist(),
                                                    format_func=lambda x: "Todos" if x==0 else clientes_df[clientes_df['id']==x]['nombre'].iloc[0],
                                                    key="filtro_cliente_cursos")
                if cliente_para_cursos != 0:
                    nombre_cliente = clientes_df[clientes_df['id']==cliente_para_cursos]['nombre'].iloc[0]
                    logo_filename = clientes_df[clientes_df['id']==cliente_para_cursos]['logo_filename'].iloc[0]
                    if pd.isna(logo_filename):
                        logo_filename = None
                    mostrar_logo_grande(nombre_cliente, size=120, custom_logo=logo_filename)
                    st.markdown(f"<h4 style='text-align: center;'>{nombre_cliente}</h4>", unsafe_allow_html=True)
                    st.markdown("---")
                if cliente_para_cursos == 0:
                    cursos_df = db.obtener_cursos_por_cliente()
                else:
                    cursos_df = db.obtener_cursos_por_cliente(cliente_para_cursos)
                if not cursos_df.empty:
                    cursos_mostrar = cursos_df.copy()
                    cursos_mostrar['vigencia_mostrar'] = cursos_mostrar['vigencia_dias'].apply(
                        lambda x: "Sin vencimiento" if x == -1 else str(x)
                    )
                    # ---------- NUEVO: Mostrar URL como enlace ----------
                    st.dataframe(
                        cursos_mostrar[['nombre', 'cliente', 'descripcion', 'duracion_horas', 'vigencia_mostrar', 'url']],
                        width='stretch',
                        hide_index=True,
                        column_config={
                            "vigencia_mostrar": "Vigencia (días)",
                            "url": st.column_config.LinkColumn("Acceso al curso")
                        }
                    )
                else:
                    st.info("No hay cursos para este cliente")
                
                # Agregar curso
                with st.expander("➕ Agregar curso"):
                    nombre_curso = st.text_input("Nombre del curso", key="nuevo_curso_nombre")
                    cliente_curso = st.selectbox("Cliente", clientes_df['id'].tolist(),
                                                format_func=lambda x: clientes_df[clientes_df['id']==x]['nombre'].iloc[0],
                                                key="nuevo_curso_cliente")
                    cliente_nombre = clientes_df[clientes_df['id']==cliente_curso]['nombre'].iloc[0]
                    cliente_logo = clientes_df[clientes_df['id']==cliente_curso]['logo_filename'].iloc[0]
                    if pd.isna(cliente_logo):
                        cliente_logo = None
                    st.markdown("**Logo del cliente seleccionado:**")
                    mostrar_logo_cliente(cliente_nombre, width=80, centered=True, custom_logo=cliente_logo)
                    desc_curso = st.text_area("Descripción", key="nuevo_curso_desc")
                    duracion = st.number_input("Duración (horas)", min_value=0, step=1, key="nuevo_curso_duracion")
                    
                    # ---------- NUEVO: Campo URL ----------
                    url_curso = st.text_input("URL del curso (opcional)", key="nuevo_curso_url", placeholder="https://ejemplo.com")
                    
                    # Checkbox para "Sin vencimiento"
                    sin_vencimiento = st.checkbox("Sin vencimiento", key="sin_vencimiento_agregar")
                    if sin_vencimiento:
                        vigencia = -1
                        st.number_input("Vigencia (días)", min_value=0, step=1, key="nuevo_curso_vigencia", disabled=True, value=0)
                    else:
                        vigencia = st.number_input("Vigencia (días)", min_value=0, step=1, key="nuevo_curso_vigencia")
                    
                    if st.button("Crear curso", key="crear_curso"):
                        if nombre_curso and cliente_curso:
                            db.agregar_curso(nombre_curso, cliente_curso, desc_curso, duracion, vigencia, url_curso)  # <-- NUEVO
                            st.success("Curso agregado")
                            st.rerun()
                        else:
                            st.warning("Nombre y cliente son obligatorios")
                
                # Editar/Eliminar curso
                with st.expander("✏️ Editar/Eliminar curso"):
                    if not cursos_df.empty:
                        curso_seleccionado = st.selectbox("Curso", cursos_df['id'].tolist(),
                                                          format_func=lambda x: f"{cursos_df[cursos_df['id']==x]['nombre'].iloc[0]} ({cursos_df[cursos_df['id']==x]['cliente'].iloc[0]})",
                                                          key="curso_seleccionado_edit")
                        if curso_seleccionado:
                            curso_row = cursos_df[cursos_df['id'] == curso_seleccionado].iloc[0]
                            cliente_logo = None
                            if not clientes_df[clientes_df['nombre']==curso_row['cliente']].empty:
                                cliente_logo = clientes_df[clientes_df['nombre']==curso_row['cliente']]['logo_filename'].iloc[0]
                                if pd.isna(cliente_logo):
                                    cliente_logo = None
                            st.markdown("**Logo del cliente asociado:**")
                            mostrar_logo_cliente(curso_row['cliente'], width=80, centered=True, custom_logo=cliente_logo)
                            
                            ed_nombre_curso = st.text_input("Nombre curso", value=curso_row['nombre'], key=f"edit_curso_nombre_{curso_seleccionado}")
                            cliente_actual = curso_row['cliente']
                            cliente_id_actual = clientes_df[clientes_df['nombre']==cliente_actual]['id'].iloc[0] if cliente_actual in clientes_df['nombre'].values else clientes_df['id'].iloc[0]
                            ed_cliente_curso = st.selectbox("Cliente", clientes_df['id'].tolist(),
                                                            index=clientes_df[clientes_df['id']==cliente_id_actual].index[0],
                                                            format_func=lambda x: clientes_df[clientes_df['id']==x]['nombre'].iloc[0],
                                                            key=f"edit_curso_cliente_{curso_seleccionado}")
                            ed_desc = st.text_area("Descripción", value=curso_row['descripcion'] if curso_row['descripcion'] else "", key=f"edit_curso_desc_{curso_seleccionado}")
                            ed_duracion = st.number_input("Duración (horas)", value=int(curso_row['duracion_horas']) if curso_row['duracion_horas'] else 0, step=1, key=f"edit_curso_duracion_{curso_seleccionado}")
                            
                            # ---------- NUEVO: Campo URL en edición ----------
                            ed_url = st.text_input("URL del curso", value=curso_row['url'] if curso_row['url'] else "", key=f"edit_curso_url_{curso_seleccionado}")
                            
                            # Manejo de vigencia con checkbox
                            vigencia_actual = curso_row['vigencia_dias']
                            es_sin_vencimiento = vigencia_actual == -1
                            sin_vencimiento_edit = st.checkbox("Sin vencimiento", value=es_sin_vencimiento, key=f"sin_vencimiento_edit_{curso_seleccionado}")
                            if sin_vencimiento_edit:
                                ed_vigencia = -1
                                st.number_input("Vigencia (días)", value=0, step=1, key=f"edit_curso_vigencia_{curso_seleccionado}", disabled=True)
                            else:
                                ed_vigencia = st.number_input("Vigencia (días)", value=int(vigencia_actual) if vigencia_actual != -1 else 0, step=1, key=f"edit_curso_vigencia_{curso_seleccionado}")
                            
                            col_cu, col_cd = st.columns(2)
                            with col_cu:
                                if st.button("Actualizar curso", key=f"actualizar_curso_{curso_seleccionado}"):
                                    db.actualizar_curso(curso_seleccionado, ed_nombre_curso, ed_cliente_curso, ed_desc, ed_duracion, ed_vigencia, ed_url)  # <-- NUEVO
                                    st.success("Curso actualizado")
                                    st.rerun()
                            with col_cd:
                                if st.button("Eliminar curso", type="primary", key=f"eliminar_curso_{curso_seleccionado}"):
                                    db.eliminar_curso(curso_seleccionado)
                                    st.success("Curso eliminado")
                                    st.rerun()
            else:
                st.warning("Primero debe crear clientes")

    # --------------------- ASIGNAR CAPACITACIONES ---------------------

    with tabs[2]:
        st.subheader("Asignación de Cursos a Empleados")
        empleados_df = db.obtener_empleados()
        cursos_df = db.obtener_cursos_por_cliente()
        if empleados_df.empty or cursos_df.empty:
            st.warning("Debe haber empleados y cursos para asignar")
        else:
            empleados_df = empleados_df.sort_values('cedula')

            # ----- Filtro para empleados -----
            if 'limpiar_filtro_empleado' not in st.session_state:
                st.session_state.limpiar_filtro_empleado = False

            if st.session_state.limpiar_filtro_empleado:
                st.session_state.filtro_cedula_empleado_asignar = ""
                st.session_state.limpiar_filtro_empleado = False

            col_filtro_emp1, col_filtro_emp2 = st.columns([3, 1])
            with col_filtro_emp1:
                filtro_cedula_empleado = st.text_input(
                    "🔍 Filtrar empleado por cédula",
                    placeholder="Ej: 1051635722",
                    key="filtro_cedula_empleado_asignar"
                )
            with col_filtro_emp2:
                if st.button("Limpiar filtro", key="limpiar_filtro_empleado_asignar"):
                    st.session_state.limpiar_filtro_empleado = True
                    st.rerun()

            filtro_actual_emp = st.session_state.get("filtro_cedula_empleado_asignar", "")
            empleados_filtrados = empleados_df.copy()
            if filtro_actual_emp:
                empleados_filtrados = empleados_filtrados[
                    empleados_filtrados['cedula'].astype(str).str.contains(filtro_actual_emp, case=False, na=False)
                ]

            if empleados_filtrados.empty:
                st.warning("No hay empleados que coincidan con el filtro.")
            else:
                col_asig1, col_asig2 = st.columns(2)
                with col_asig1:
                    empleado_asig = st.selectbox(
                        "Empleado",
                        empleados_filtrados['id'].tolist(),
                        format_func=lambda x: f"{empleados_filtrados[empleados_filtrados['id']==x]['cedula'].iloc[0]} - {empleados_filtrados[empleados_filtrados['id']==x]['nombre'].iloc[0]}",
                        key="asignar_empleado"
                    )
                with col_asig2:
                    curso_asig = st.selectbox(
                        "Curso",
                        cursos_df['id'].tolist(),
                        format_func=lambda x: f"{cursos_df[cursos_df['id']==x]['nombre'].iloc[0]} ({cursos_df[cursos_df['id']==x]['cliente'].iloc[0]})",
                        key="asignar_curso"
                    )
                    # Obtener información del curso seleccionado
                    curso_seleccionado = cursos_df[cursos_df['id'] == curso_asig]
                    if not curso_seleccionado.empty:
                        vigencia = curso_seleccionado.iloc[0]['vigencia_dias']
                        # Convertir a int de Python para timedelta
                        if isinstance(vigencia, (int, float, np.integer, np.floating)):
                            vigencia = int(vigencia)
                        else:
                            vigencia = 365
                        cliente_curso = curso_seleccionado.iloc[0]['cliente']
                        # Mostrar logo del cliente
                        cliente_logo = None
                        if not clientes_df[clientes_df['nombre']==cliente_curso].empty:
                            cliente_logo = clientes_df[clientes_df['nombre']==cliente_curso]['logo_filename'].iloc[0]
                            if pd.isna(cliente_logo):
                                cliente_logo = None
                        mostrar_logo_cliente(cliente_curso, width=60, centered=True, custom_logo=cliente_logo)

                # Fecha de asignación (editable, por defecto hoy)
                fecha_asig = st.date_input(
                    "📅 Fecha de asignación",
                    value=date.today(),
                    key="fecha_asignacion_asig",
                    help="Fecha en que se asigna el curso"
                )

                # --- Cálculo automático de la fecha de vencimiento ---
                if vigencia == -1:
                    st.info("📌 Este curso no tiene vencimiento (vigencia indefinida).")
                    fecha_vencimiento_texto = None
                    # No mostramos campo de vencimiento
                else:
                    fecha_vencimiento_calc = fecha_asig + timedelta(days=vigencia)
                    fecha_vencimiento_texto = fecha_vencimiento_calc.isoformat()
                    # Mostrar la fecha calculada como texto (se actualiza automáticamente al cambiar la fecha de asignación)
                    st.write(f"📅 **Fecha de vencimiento (calculada automáticamente):** {fecha_vencimiento_calc.strftime('%d/%m/%Y')}")

                    # --- BOTÓN DE ASIGNACIÓN CON VALIDACIÓN ---

                sin_ejecucion = st.checkbox("🔄 Sin ejecución (asignado pero no iniciado)", key="sin_ejecucion_check")
                estado_asignacion = 'sin_ejecucion' if sin_ejecucion else 'pendiente'

                if st.button("Asignar curso", key="btn_asignar"):
                    if db.empleado_tiene_curso_vigente(empleado_asig, curso_asig):
                        st.session_state.mensaje_asignacion = ("error", "❌ El empleado ya tiene este curso en estado vigente. No se puede asignar de nuevo.")
                    else:
                        db.asignar_curso(
                            empleado_asig,
                            curso_asig,
                            fecha_vencimiento=fecha_vencimiento_texto,
                            fecha_asignacion=fecha_asig.isoformat(),
                            estado=estado_asignacion
                        )
                        st.session_state.mensaje_asignacion = ("exito", "✅ Curso asignado correctamente.")
                    st.rerun()

                # Mostrar mensaje almacenado
                if 'mensaje_asignacion' in st.session_state:
                    tipo, texto = st.session_state.mensaje_asignacion
                    if tipo == "exito":
                        st.success(texto)
                    else:
                        st.error(texto)
                    del st.session_state.mensaje_asignacion

            st.markdown("---")
            st.subheader("Capacitaciones asignadas")
            asignaciones = db.obtener_asignaciones()
            if not asignaciones.empty:
                # ----- Filtro para la tabla de asignaciones -----
                if 'limpiar_filtro_cedula_bandera' not in st.session_state:
                    st.session_state.limpiar_filtro_cedula_bandera = False

                if st.session_state.limpiar_filtro_cedula_bandera:
                    st.session_state.input_filtro_cedula = ""
                    st.session_state.limpiar_filtro_cedula_bandera = False

                col_filtro1, col_filtro2 = st.columns([3, 1])
                with col_filtro1:
                    filtro_cedula = st.text_input(
                        "🔍 Filtrar por cédula (parcial o completa)",
                        placeholder="Ej: 1051635722",
                        key="input_filtro_cedula"
                    )
                with col_filtro2:
                    if st.button("Limpiar filtro", key="limpiar_filtro_cedula"):
                        st.session_state.limpiar_filtro_cedula_bandera = True
                        st.rerun()

                filtro_actual = st.session_state.get("input_filtro_cedula", "")
                if filtro_actual:
                    asignaciones = asignaciones[asignaciones['cedula'].astype(str).str.contains(filtro_actual, case=False, na=False)]

                if asignaciones.empty:
                    st.info("No hay asignaciones que coincidan con el filtro.")
                else:
                    # --- DataFrame principal con colores ---
                    df_display = asignaciones.copy()
                    df_display['fecha_asignacion'] = pd.to_datetime(df_display['fecha_asignacion'])
                    df_display['fecha_vencimiento'] = pd.to_datetime(df_display['fecha_vencimiento'])

                    hoy = pd.Timestamp(date.today())
                    limite = hoy + pd.Timedelta(days=30)

                    def estado_logico(row):
                        """Devuelve el estado de vigencia según la fecha de vencimiento."""
                        if row['fecha_vencimiento'].year >= 9999:
                            return "Sin vencimiento"
                        if row['fecha_vencimiento'] < hoy:
                            return "Vencido"
                        if row['fecha_vencimiento'] <= limite:
                            return "Próximo a vencer"
                        return "Vigente"

                    def color_estado(estado):
                        if estado == "Vigente":
                            return "background-color: #d4edda; color: #155724;"
                        elif estado == "Próximo a vencer":
                            return "background-color: #fff3cd; color: #856404;"
                        elif estado == "Vencido":
                            return "background-color: #f8d7da; color: #721c24;"
                        elif estado == "Sin vencimiento":
                            return "background-color: #d1ecf1; color: #155724;"
                        else:
                            return ""

                    df_display['estado_mostrar'] = df_display.apply(estado_logico, axis=1)
                    df_display['color'] = df_display['estado_mostrar'].apply(color_estado)

                    styled_df = df_display[['empleado', 'cedula', 'curso', 'cliente', 'fecha_asignacion', 'fecha_vencimiento', 'estado_mostrar']].style.apply(
                        lambda x: [df_display.loc[x.name, 'color'] for _ in x] if x.name in df_display.index else [''] * len(x),
                        axis=1
                    )

                    st.dataframe(
                        styled_df,
                        column_config={
                            "fecha_asignacion": st.column_config.DateColumn("Fecha asignación", format="DD/MM/YYYY"),
                            "fecha_vencimiento": st.column_config.DateColumn("Fecha vencimiento", format="DD/MM/YYYY"),
                            "cedula": "Cédula",
                            "estado_mostrar": "Estado"
                        },
                        width='stretch',
                        height=400
                    )

                    # ----- EDITOR INTERACTIVO CON FILTRO INTERNO, FECHA ASIGNACIÓN EDITABLE Y ESTADOS LÓGICOS -----
                    with st.expander("✏️ Editar estados o eliminar asignaciones", expanded=False):
                        st.info("Modifica el estado (vigente, próximo a vencer, vencido) o la fecha de asignación/vencimiento y haz clic en 'Guardar cambios'. Para eliminar, marca el checkbox.")

                        # Preparar dataframe editable
                        df_edit = asignaciones[['id', 'empleado', 'cedula', 'curso', 'cliente', 'estado']].copy()
                        df_edit['fecha_asignacion'] = pd.to_datetime(asignaciones['fecha_asignacion'])
                        df_edit['fecha_vencimiento'] = pd.to_datetime(asignaciones['fecha_vencimiento'])

                        # Excluir sin vencimiento
                        df_edit_no_editable = df_edit[df_edit['fecha_vencimiento'].dt.year >= 9999].copy()
                        df_edit_editable = df_edit[df_edit['fecha_vencimiento'].dt.year < 9999].copy()

                        if not df_edit_no_editable.empty:
                            st.warning(f"{len(df_edit_no_editable)} asignaciones tienen 'Sin vencimiento' y no se pueden editar desde aquí.")

                        if df_edit_editable.empty:
                            st.info("No hay asignaciones con vencimiento definido para editar.")
                        else:
                            # Convertir estado BD a estado lógico para mostrar en el editor
                            df_edit_editable['estado_logico'] = df_edit_editable.apply(estado_logico, axis=1)
                            # Asegurar que solo sean los tres posibles (vigente, próximo, vencido)
                            df_edit_editable['estado_logico'] = df_edit_editable['estado_logico'].apply(
                                lambda x: x if x in ['Vigente', 'Próximo a vencer', 'Vencido', 'Sin ejecución' ] else 'Vigente'
                            )

                            # ----- FILTRO INTERNO POR CÉDULA -----
                            if 'filtro_cedula_editor_valor' not in st.session_state:
                                st.session_state.filtro_cedula_editor_valor = ''

                            col_filtro_editor1, col_filtro_editor2 = st.columns([3, 1])
                            with col_filtro_editor1:
                                filtro_cedula_editor = st.text_input(
                                    "🔍 Filtrar por cédula (dentro del editor)",
                                    placeholder="Ej: 1051635722",
                                    value=st.session_state.filtro_cedula_editor_valor
                                )
                            with col_filtro_editor2:
                                if st.button("Limpiar filtro", key="limpiar_filtro_editor"):
                                    st.session_state.filtro_cedula_editor_valor = ""
                                    st.rerun()

                            # --- SELECCIÓN EXPLÍCITA DE COLUMNAS (EXCLUYENDO 'estado') ---
                            columnas_editor = ['id', 'empleado', 'cedula', 'curso', 'cliente',
                                            'fecha_asignacion', 'fecha_vencimiento', 'estado_logico']
                            df_filtrado = df_edit_editable[columnas_editor].copy()
                            if filtro_cedula_editor:
                                df_filtrado = df_filtrado[
                                    df_filtrado['cedula'].astype(str).str.contains(filtro_cedula_editor, case=False, na=False)
                                ]

                            if df_filtrado.empty:
                                st.info("No hay asignaciones que coincidan con el filtro interno.")
                            else:
                                # Agregar columna para eliminar
                                df_filtrado['Eliminar'] = False

                                column_config_editable = {
                                    "id": None,
                                    "empleado": "Empleado",
                                    "cedula": "Cédula",
                                    "curso": "Curso",
                                    "cliente": "Cliente",
                                    "fecha_asignacion": st.column_config.DateColumn("Fecha asignación", format="DD/MM/YYYY", disabled=False),
                                    "fecha_vencimiento": st.column_config.DateColumn("Fecha vencimiento", format="DD/MM/YYYY", disabled=False),
                                    "estado_logico": st.column_config.SelectboxColumn(
                                        "Estado",
                                        options=['Vigente', 'Próximo a vencer', 'Vencido', 'Sin ejecución', 'Completado'],
                                        required=True
                                    ),
                                    "Eliminar": st.column_config.CheckboxColumn("Eliminar", default=False)
                                }

                                edited_df = st.data_editor(
                                    df_filtrado,
                                    column_config=column_config_editable,
                                    hide_index=True,
                                    width='stretch',
                                    key="editor_asignaciones"
                                )

                                col_guardar, col_eliminar = st.columns(2)
                                with col_guardar:
                                    if st.button("💾 Guardar cambios", key="guardar_estados"):
                                        cambios_estado = []
                                        cambios_fecha_asignacion = []
                                        cambios_fecha_vencimiento = []
                                        for idx, row in edited_df.iterrows():
                                            original = df_edit_editable[df_edit_editable['id'] == row['id']]
                                            if not original.empty:
                                                orig = original.iloc[0]
                                                # Comparar estado lógico
                                                if orig['estado_logico'] != row['estado_logico']:
                                                    # Mapear a estado BD
                                                    mapeo = {
                                                        'Vigente': 'pendiente',
                                                        'Próximo a vencer': 'pendiente',
                                                        'Vencido': 'vencido',
                                                        'Sin ejecución': 'sin_ejecucion',
                                                        'Completado': 'completado'
                                                    }
                                                    nuevo_estado_bd = mapeo[row['estado_logico']]
                                                    cambios_estado.append((row['id'], nuevo_estado_bd))
                                                # Comparar fecha asignación
                                                if orig['fecha_asignacion'].date() != row['fecha_asignacion'].date():
                                                    cambios_fecha_asignacion.append((row['id'], row['fecha_asignacion']))
                                                # Comparar fecha vencimiento
                                                if orig['fecha_vencimiento'].date() != row['fecha_vencimiento'].date():
                                                    cambios_fecha_vencimiento.append((row['id'], row['fecha_vencimiento']))

                                        if cambios_estado or cambios_fecha_asignacion or cambios_fecha_vencimiento:
                                            # Actualizar usando la nueva función
                                            for id_asig, nuevo_estado in cambios_estado:
                                                # Buscar si también cambió fecha_vencimiento o fecha_asignacion
                                                fecha_venc = next((f for (i, f) in cambios_fecha_vencimiento if i == id_asig), None)
                                                fecha_asig = next((f for (i, f) in cambios_fecha_asignacion if i == id_asig), None)
                                                if fecha_venc is not None and fecha_asig is not None:
                                                    db.actualizar_asignacion(id_asig, estado=nuevo_estado,
                                                                            fecha_vencimiento=fecha_venc.strftime("%Y-%m-%d"),
                                                                            fecha_asignacion=fecha_asig.strftime("%Y-%m-%d"))
                                                elif fecha_venc is not None:
                                                    db.actualizar_asignacion(id_asig, estado=nuevo_estado,
                                                                            fecha_vencimiento=fecha_venc.strftime("%Y-%m-%d"))
                                                elif fecha_asig is not None:
                                                    db.actualizar_asignacion(id_asig, estado=nuevo_estado,
                                                                            fecha_asignacion=fecha_asig.strftime("%Y-%m-%d"))
                                                else:
                                                    db.actualizar_asignacion(id_asig, estado=nuevo_estado)
                                            # Actualizar fechas que no tuvieron cambio de estado
                                            for id_asig, nueva_fecha in cambios_fecha_vencimiento:
                                                if not any(id_asig == i for (i, _) in cambios_estado):
                                                    db.actualizar_asignacion(id_asig, fecha_vencimiento=nueva_fecha.strftime("%Y-%m-%d"))
                                            for id_asig, nueva_fecha in cambios_fecha_asignacion:
                                                if not any(id_asig == i for (i, _) in cambios_estado):
                                                    db.actualizar_asignacion(id_asig, fecha_asignacion=nueva_fecha.strftime("%Y-%m-%d"))

                                            total = len(cambios_estado) + len([f for f in cambios_fecha_vencimiento if not any(f[0] == i for i, _ in cambios_estado)]) + len([f for f in cambios_fecha_asignacion if not any(f[0] == i for i, _ in cambios_estado)])
                                            st.success(f"✅ {total} cambios guardados correctamente")
                                            st.rerun()
                                        else:
                                            st.info("No se detectaron cambios.")

                                with col_eliminar:
                                    if st.button("🗑️ Eliminar seleccionadas", key="eliminar_seleccionadas"):
                                        to_delete = edited_df[edited_df['Eliminar'] == True]
                                        if not to_delete.empty:
                                            for _, row in to_delete.iterrows():
                                                db.eliminar_asignacion(row['id'])
                                            st.success(f"✅ {len(to_delete)} asignaciones eliminadas correctamente")
                                            st.rerun()
                                        else:
                                            st.warning("No hay filas marcadas para eliminar.")
            else:
                st.info("No hay asignaciones aún")

        # --------------------- INDICADORES -------------------
    with tabs[3]:
        st.subheader("📈 Tablero de Indicadores de Cumplimiento")

        # Cargar datos
        indicadores_df  = db.obtener_indicadores_resumen()
        asignaciones_df = db.obtener_asignaciones()

        if not indicadores_df.empty and not asignaciones_df.empty:
            # ── Preparar fechas y estado lógico ─────────────────────────────
            asignaciones_df['fecha_asignacion']  = pd.to_datetime(asignaciones_df['fecha_asignacion'])
            asignaciones_df['fecha_vencimiento'] = pd.to_datetime(asignaciones_df['fecha_vencimiento'])
            hoy = pd.Timestamp(date.today())

            # Mapeo de categorías internas (db.clasificar_estado) a etiquetas de UI
            MAPEO_ESTADO_DISPLAY = {
                'completado':      'Completado',
                'sin_ejecucion':   'Sin ejecución',
                'sin_vencimiento': 'Sin vencimiento',
                'vencido':         'Vencido',
                'proximo':         'Próximo a vencer',
                'vigente':         'Vigente',
            }

            def estado_logico(row):
                cat = db.clasificar_estado(row['fecha_vencimiento'], hoy)
                return MAPEO_ESTADO_DISPLAY[cat]

            asignaciones_df['estado_logico'] = asignaciones_df.apply(estado_logico, axis=1)
            asignaciones_df['necesita_renovacion'] = asignaciones_df.apply(
                lambda row: db.necesita_renovacion(row['estado'], row['fecha_vencimiento'], hoy),
                axis=1
            )

            # ── Cálculos base ────────────────────────────────────────────────
            total_empleados              = len(db.obtener_empleados())
            empleados_con_asignaciones   = len(indicadores_df[indicadores_df['total'] > 0])
            empleados_sin_asignacion     = total_empleados - empleados_con_asignaciones

            total_asignados        = indicadores_df['total'].sum()
            total_completados      = (asignaciones_df['estado'] == 'completado').sum()
            total_vencidos          = (asignaciones_df['estado_logico'] == 'Vencido').sum()
            total_proximos          = (asignaciones_df['estado_logico'] == 'Próximo a vencer').sum()
            total_vigentes          = (asignaciones_df['estado_logico'] == 'Vigente').sum()
            total_sin_ejecucion     = (asignaciones_df['estado'] == 'sin_ejecucion').sum()
            total_sin_vencimiento   = (asignaciones_df['estado_logico'] == 'Sin vencimiento').sum()
            total_por_renovar       = asignaciones_df['necesita_renovacion'].sum()

            # ── Porcentajes ──────────────────────────────────────────────────
            # Agregado: peso proporcional a la carga de cada empleado
            pct_agregado   = (total_completados / total_asignados * 100) if total_asignados > 0 else 0
            # Individual: cada empleado pesa igual
            pct_individual = indicadores_df['porcentaje'].mean() if not indicadores_df.empty else 0
            # Riesgo activo: cursos ya vencidos
            pct_vencidos   = (total_vencidos / total_asignados * 100) if total_asignados > 0 else 0
            # Riesgo próximo: vencen en 30 días
            pct_proximos   = (total_proximos / total_asignados * 100) if total_asignados > 0 else 0
            # Exposición total: vencidos + próximos
            pct_exposicion = pct_vencidos + pct_proximos

            # ── Fila 1: volúmenes ────────────────────────────────────────────
            st.markdown("##### Resumen general")
            c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
            c1.metric("👥 Empleados",          total_empleados,
                    delta=f"{empleados_sin_asignacion} sin asignar" if empleados_sin_asignacion > 0 else "Todos asignados",
                    delta_color="inverse" if empleados_sin_asignacion > 0 else "normal")
            c2.metric("📚 Asignados",          total_asignados)
            c3.metric("✅ Completados",         total_completados)
            c4.metric("🟢 Vigentes",           total_vigentes)
            c5.metric("⚠️ Próximos a vencer",  total_proximos)
            c6.metric("🔴 Vencidos",           total_vencidos)
            c7.metric("♾️ Sin vencimiento",    total_sin_vencimiento)
            c8.metric("🔄 Por renovar (≤30d)", int(total_por_renovar),
                    help="Cursos ya completados cuya vigencia expira pronto y requieren reasignación. No afecta el % de cumplimiento.")

            # ── Detalle de cursos completados que requieren renovación ────────
            if total_por_renovar > 0:
                with st.expander(f"🔄 Ver detalle de los {int(total_por_renovar)} cursos por renovar", expanded=False):
                    df_renovar = asignaciones_df[asignaciones_df['necesita_renovacion']].copy()
                    df_renovar['dias_restantes'] = (df_renovar['fecha_vencimiento'] - hoy).dt.days
                    df_renovar = df_renovar.sort_values('dias_restantes')

                    # Filtro opcional por empleado o cliente
                    col_fr1, col_fr2 = st.columns(2)
                    with col_fr1:
                        clientes_renovar = ["Todos"] + sorted(df_renovar['cliente'].unique().tolist())
                        cliente_filtro_renovar = st.selectbox("Filtrar por cliente", clientes_renovar, key="filtro_cliente_renovar")
                    with col_fr2:
                        filtro_emp_renovar = st.text_input("Buscar por nombre o cédula", key="filtro_emp_renovar")

                    df_renovar_filtrado = df_renovar.copy()
                    if cliente_filtro_renovar != "Todos":
                        df_renovar_filtrado = df_renovar_filtrado[df_renovar_filtrado['cliente'] == cliente_filtro_renovar]
                    if filtro_emp_renovar:
                        df_renovar_filtrado = df_renovar_filtrado[
                            df_renovar_filtrado['empleado'].str.contains(filtro_emp_renovar, case=False, na=False) |
                            df_renovar_filtrado['cedula'].astype(str).str.contains(filtro_emp_renovar, case=False, na=False)
                        ]

                    if df_renovar_filtrado.empty:
                        st.info("No hay resultados para este filtro.")
                    else:
                        # Colorear por urgencia
                        def color_renovacion(dias):
                            if dias <= 7:
                                return "background-color: #f8d7da; color: #721c24;"
                            elif dias <= 15:
                                return "background-color: #fff3cd; color: #856404;"
                            else:
                                return "background-color: #d1ecf1; color: #0c5460;"

                        df_renovar_filtrado['color'] = df_renovar_filtrado['dias_restantes'].apply(color_renovacion)
                        df_renovar_filtrado['fecha_vencimiento_mostrar'] = df_renovar_filtrado['fecha_vencimiento'].dt.strftime("%d/%m/%Y")

                        styled_renovar = df_renovar_filtrado[
                            ['empleado', 'cedula', 'curso', 'cliente', 'fecha_vencimiento_mostrar', 'dias_restantes']
                        ].style.apply(
                            lambda x: [df_renovar_filtrado.loc[x.name, 'color'] for _ in x],
                            axis=1
                        )

                        st.dataframe(
                            styled_renovar,
                            column_config={
                                "empleado": "Empleado",
                                "cedula": "Cédula",
                                "curso": "Curso",
                                "cliente": "Cliente",
                                "fecha_vencimiento_mostrar": "Vence el",
                                "dias_restantes": st.column_config.NumberColumn("Días restantes", format="%d")
                            },
                            hide_index=True,
                            width='stretch'
                        )

                        # Descarga en Excel
                        import io
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                        def generar_excel_renovaciones(df):
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Por Renovar"

                            header_font = Font(bold=True, color="FFFFFF")
                            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                            center_align = Alignment(horizontal="center", vertical="center")

                            headers = ['Empleado', 'Cédula', 'Curso', 'Cliente', 'Vence el', 'Días restantes']
                            for col_idx, header in enumerate(headers, 1):
                                cell = ws.cell(row=1, column=col_idx, value=header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_align
                                cell.border = border

                            for r_idx, row in df.iterrows():
                                row_num = r_idx + 2 if isinstance(r_idx, int) else list(df.index).index(r_idx) + 2
                            for i, (_, row) in enumerate(df.iterrows(), 2):
                                ws.cell(row=i, column=1, value=row['empleado']).border = border
                                ws.cell(row=i, column=2, value=row['cedula']).border = border
                                ws.cell(row=i, column=3, value=row['curso']).border = border
                                ws.cell(row=i, column=4, value=row['cliente']).border = border
                                ws.cell(row=i, column=5, value=row['fecha_vencimiento_mostrar']).border = border
                                dias_cell = ws.cell(row=i, column=6, value=row['dias_restantes'])
                                dias_cell.border = border
                                dias_cell.alignment = center_align
                                if row['dias_restantes'] <= 7:
                                    dias_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                elif row['dias_restantes'] <= 15:
                                    dias_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                                else:
                                    dias_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")

                            ws.column_dimensions['A'].width = 30
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 30
                            ws.column_dimensions['D'].width = 20
                            ws.column_dimensions['E'].width = 15
                            ws.column_dimensions['F'].width = 15

                            wb.save(output)
                            output.seek(0)
                            return output

                        excel_renovar = generar_excel_renovaciones(df_renovar_filtrado)
                        st.download_button(
                            label="📥 Descargar renovaciones pendientes en Excel",
                            data=excel_renovar,
                            file_name=f'renovaciones_pendientes_{date.today().strftime("%Y%m%d")}.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='download_renovaciones_excel'
                        )                            

            st.markdown("##### Indicadores de cumplimiento")

            # ── Fila 2: porcentajes con contexto ────────────────────────────
            p1, p2, p3, p4, p5 = st.columns(5)
            p1.metric(
                "📊 Cumplimiento agregado",
                f"{pct_agregado:.1f}%",
                help="Completados / Asignados totales. Empleados con más cursos pesan más.",
            )
            p2.metric(
                "👤 Cumplimiento por persona",
                f"{pct_individual:.1f}%",
                help="Promedio de los porcentajes individuales. Cada empleado pesa igual.",
            )
            p3.metric(
                "🔴 % Vencidos",
                f"{pct_vencidos:.1f}%",
                help="Cursos vencidos sin completar sobre el total asignado.",
                delta=f"-{pct_vencidos:.1f}% riesgo activo",
                delta_color="inverse",
            )
            p4.metric(
                "⚠️ % Próximos",
                f"{pct_proximos:.1f}%",
                help="Cursos que vencen en los próximos 30 días sobre el total asignado.",
                delta=f"-{pct_proximos:.1f}% riesgo potencial",
                delta_color="inverse",
            )
            p5.metric(
                "🚨 Exposición total",
                f"{pct_exposicion:.1f}%",
                help="Vencidos + Próximos sobre el total asignado. Riesgo combinado.",
                delta=f"-{pct_exposicion:.1f}% cursos en riesgo",
                delta_color="inverse",
            )

            # ── Alerta si los dos cumplimientos difieren mucho ───────────────
            if abs(pct_agregado - pct_individual) >= 10:
                st.warning(
                    f"⚠️ El cumplimiento agregado ({pct_agregado:.1f}%) y por persona ({pct_individual:.1f}%) "
                    f"difieren en {abs(pct_agregado - pct_individual):.1f} puntos. "
                    "Hay empleados con cargas de cursos muy diferentes entre sí — "
                    "revise la distribución de asignaciones."
                )

            st.markdown("---")
            
            # ----- 1. Panorama general (Donut + Barras) -----
            st.subheader("📊 Panorama General de Capacitaciones")
            
            # ---- Donut: distribución de estados ----
            col1, col2 = st.columns([1, 2])
            
            with col1:
                estado_counts = asignaciones_df['estado_logico'].value_counts().reset_index()
                estado_counts.columns = ['Estado', 'Cantidad']
                colores_donut = {
                    'Vigente': '#28a745',
                    'Próximo a vencer': '#ffc107',
                    'Vencido': '#dc3545',
                    'Sin vencimiento': '#6c757d'
                }
                fig_donut = px.pie(
                    estado_counts,
                    values='Cantidad',
                    names='Estado',
                    color='Estado',
                    color_discrete_map=colores_donut,
                    hole=0.5,
                    title="Distribución de estados"
                )
                fig_donut.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig_donut, width='stretch')
                
                # Análisis donut
                pct_vencidos    = (total_vencidos / total_asignados * 100) if total_asignados else 0
                pct_proximos    = (total_proximos / total_asignados * 100) if total_asignados else 0
                pct_vigentes    = (total_vigentes / total_asignados * 100) if total_asignados else 0
                pct_renovar     = (total_por_renovar / total_asignados * 100) if total_asignados else 0
                st.markdown(f"""
                **📌 Análisis:**  
                - **{pct_vigentes:.1f}%** están vigentes (fecha no vencida).  
                - **{pct_vencidos:.1f}%** están vencidas → **requieren atención inmediata**.  
                - **{pct_proximos:.1f}%** vencen en los próximos 30 días → **acción preventiva**.  
                - **{pct_renovar:.1f}%** de las completadas requieren renovación → **programar reasignación**.
                """)
            
            with col2:
                # ---- Barras de cumplimiento con TODOS los empleados ----
                empleados_todos = db.obtener_empleados()
                import unicodedata
                def normalizar_nombre(nombre):
                    nombre = nombre.strip().upper()
                    nombre = unicodedata.normalize('NFKD', nombre).encode('ASCII', 'ignore').decode('ASCII')
                    nombre = ' '.join(nombre.split())
                    return nombre
                
                empleados_todos['nombre_norm'] = empleados_todos['nombre'].apply(normalizar_nombre)
                indicadores_df['nombre_norm'] = indicadores_df['nombre'].apply(normalizar_nombre)
                
                indicadores_con_cedula = indicadores_df.merge(
                    empleados_todos[['nombre_norm', 'cedula']],
                    on='nombre_norm',
                    how='left'
                )
                pct_dict = dict(zip(indicadores_con_cedula['cedula'], indicadores_con_cedula['porcentaje']))
                
                datos = []
                for _, row in empleados_todos.iterrows():
                    cedula = row['cedula']
                    nombre = row['nombre']
                    pct = pct_dict.get(cedula, 0.0)
                    datos.append({'nombre': nombre, 'cedula': cedula, 'porcentaje': pct})
                cumplimiento_todos = pd.DataFrame(datos)
                
                cumplimiento_emp = cumplimiento_todos.sort_values('porcentaje', ascending=True)
                altura = max(400, min(1500, len(empleados_todos) * 20))
                
                fig_bar = px.bar(
                    cumplimiento_emp,
                    x='porcentaje',
                    y='nombre',
                    orientation='h',
                    color='porcentaje',
                    color_continuous_scale=['red', 'orange', 'green'],
                    range_color=[0, 100],
                    title=f"Cumplimiento por empleado (%) - Total: {len(empleados_todos)} empleados",
                    labels={'porcentaje': 'Cumplimiento %', 'nombre': ''}
                )
                promedio = indicadores_df['porcentaje'].mean() if not indicadores_df.empty else 0
                fig_bar.add_vline(
                    x=promedio,
                    line_dash="dash",
                    line_color="blue",
                    annotation_text=f"Promedio: {promedio:.1f}%",
                    annotation_font_color="black",
                    annotation_position="top"
                )
                fig_bar.update_layout(height=altura)
                st.plotly_chart(fig_bar, width='stretch')
                
                # Análisis barras
                empleados_bajo = cumplimiento_emp[cumplimiento_emp['porcentaje'] < 50]
                empleados_alto = cumplimiento_emp[cumplimiento_emp['porcentaje'] >= 80]
                st.markdown(f"""
                **📌 Análisis:**  
                - **{len(empleados_bajo)} empleados** tienen cumplimiento **< 50%** → priorizar acciones de mejora.  
                - **{len(empleados_alto)} empleados** superan el **80%** → reconocer y compartir buenas prácticas.  
                - Cumplimiento promedio general: **{promedio:.1f}%**.
                """)
            
            # ----- 2. Radar por Cliente -----
            st.subheader("📡 Rendimiento por Cliente")
            cliente_stats = asignaciones_df.groupby('cliente').agg(
                total=('estado', 'count'),
                completados=('estado', lambda x: (x=='completado').sum())
            ).reset_index()
            cliente_stats['porcentaje'] = (cliente_stats['completados'] / cliente_stats['total'] * 100).round(1)
            
            categorias = cliente_stats['cliente'].tolist()
            valores = cliente_stats['porcentaje'].tolist()
            categorias.append(categorias[0])
            valores.append(valores[0])
            fig_radar = go.Figure(
                data=go.Scatterpolar(
                    r=valores,
                    theta=categorias,
                    fill='toself',
                    name='Cumplimiento %'
                ),
                layout=go.Layout(
                    polar=dict(
                        radialaxis=dict(
                            visible=True,
                            range=[0, 100]
                        )
                    ),
                    showlegend=False,
                    title="Porcentaje de cursos completados por cliente"
                )
            )
            st.plotly_chart(fig_radar, width='stretch')
            
            # Análisis radar
            if not cliente_stats.empty:
                mejor_cliente = cliente_stats.loc[cliente_stats['porcentaje'].idxmax()]
                peor_cliente = cliente_stats.loc[cliente_stats['porcentaje'].idxmin()]
                st.markdown(f"""
                **📌 Análisis:**  
                - Mejor desempeño: **{mejor_cliente['cliente']}** con **{mejor_cliente['porcentaje']:.1f}%** de cumplimiento.  
                - Área de oportunidad: **{peor_cliente['cliente']}** con **{peor_cliente['porcentaje']:.1f}%**.  
                - Considere reuniones específicas con el cliente de menor rendimiento para identificar barreras.
                """)
            
            # ----- 3. Evolución de vencimientos en el tiempo -----
            st.subheader("📈 Tendencia de Vencimientos")
            vencimientos = asignaciones_df[asignaciones_df['fecha_vencimiento'].dt.year < 9999].copy()
            vencimientos['mes_vencimiento'] = vencimientos['fecha_vencimiento'].dt.to_period('M').dt.start_time
            tendencia = vencimientos.groupby('mes_vencimiento').size().reset_index(name='cantidad')
            if not tendencia.empty:
                fig_line = px.line(
                    tendencia, 
                    x='mes_vencimiento', 
                    y='cantidad',
                    title='Cantidad de cursos que vencen por mes',
                    labels={'mes_vencimiento': 'Mes', 'cantidad': 'Cursos'}
                )
                st.plotly_chart(fig_line, width='stretch')
                
                # Análisis tendencia
                max_mes = tendencia.loc[tendencia['cantidad'].idxmax()]
                promedio_mensual = tendencia['cantidad'].mean()
                st.markdown(f"""
                **📌 Análisis:**  
                - Pico máximo de vencimientos en **{max_mes['mes_vencimiento'].strftime('%B %Y')}** con **{int(max_mes['cantidad'])}** cursos.  
                - Promedio mensual: **{promedio_mensual:.1f}** cursos.  
                - Planifique con anticipación los meses de alta demanda para evitar acumulación de vencimientos.
                """)
            else:
                st.info("No hay datos de vencimientos para mostrar tendencia.")
            
            # ----- 4. Clustering de empleados (K-means) -----
            st.subheader("🧠 Segmentación de Empleados por Desempeño")
            try:
                from sklearn.cluster import KMeans
                from sklearn.preprocessing import StandardScaler
                from sklearn.decomposition import PCA

                # ── 1. Preparar features ────────────────────────────────────────
                FEATURES = ['total', 'completados', 'porcentaje', 'vigentes', 'vencidos', 'proximos']
                X_scaled = StandardScaler().fit_transform(indicadores_df[FEATURES])

                # ── 2. Método del codo ──────────────────────────────────────────
                K_MIN, K_MAX = 2, 7            # rango válido para n=68
                k_list   = list(range(K_MIN, K_MAX + 1))
                inertias = []
                for k in k_list:
                    km = KMeans(n_clusters=k, random_state=42, n_init=10)
                    km.fit(X_scaled)
                    inertias.append(km.inertia_)

                fig_codo = px.line(
                    {'k': k_list, 'inercia': inertias},
                    x='k', y='inercia', markers=True,
                    title='Método del codo — selección de k óptimo',
                    labels={'k': 'Número de clusters (k)', 'inercia': 'Inercia (WCSS)'},
                    width=700, height=380,
                )
                fig_codo.update_traces(marker=dict(size=9))
                st.plotly_chart(fig_codo, width='stretch')

                # ── 3. k óptimo: punto de mayor caída de inercia ────────────────
                # La inercia siempre decrece con k, así que cada diff es positiva.
                # El codo es donde la reducción es máxima → max(diffs).
                diffs   = [inertias[i] - inertias[i + 1] for i in range(len(inertias) - 1)]
                k_opt   = k_list[diffs.index(max(diffs)) + 1]   # +1: diffs[0] compara k_list[0]→k_list[1]
                k_opt   = max(K_MIN, min(k_opt, K_MAX))          # clamp dentro del rango definido

                # Varianza explicada por PCA (calidad visual del scatter)
                pca           = PCA(n_components=2)
                componentes   = pca.fit_transform(X_scaled)
                var_total     = round(sum(pca.explained_variance_ratio_) * 100, 1)

                st.caption(
                    f"🔍 k óptimo seleccionado = **{k_opt}** "
                    f"| PCA explica el **{var_total}%** de la varianza"
                )
                if var_total < 70:
                    st.warning(
                        "⚠️ El gráfico PCA captura menos del 70 % de la varianza. "
                        "La separación visual entre clusters puede ser parcial; "
                        "confíe en la tabla de estadísticas para interpretar los grupos."
                    )

                # ── 4. Clustering final ─────────────────────────────────────────
                kmeans = KMeans(n_clusters=k_opt, random_state=42, n_init=10)
                indicadores_df['cluster']  = kmeans.fit_predict(X_scaled)
                indicadores_df['pca1']    = componentes[:, 0]
                indicadores_df['pca2']    = componentes[:, 1]

                # ── 5. Nombres descriptivos por nivel de cumplimiento ───────────
                NOMBRES_POR_K = {
                    2: ["Bajo", "Alto"],
                    3: ["Bajo", "Medio", "Alto"],
                    4: ["Bajo", "Medio-bajo", "Medio-alto", "Alto"],
                    5: ["Muy bajo", "Bajo", "Medio", "Alto", "Muy alto"],
                    6: ["Muy bajo", "Bajo", "Medio-bajo", "Medio-alto", "Alto", "Muy alto"],
                    7: ["Muy bajo", "Bajo", "Medio-bajo", "Medio", "Medio-alto", "Alto", "Muy alto"],
                }
                cluster_means = indicadores_df.groupby('cluster')['porcentaje'].mean().sort_values()
                nombres       = NOMBRES_POR_K[k_opt]
                label_map     = {cluster: nombre for nombre, (cluster, _) in zip(nombres, cluster_means.items())}
                indicadores_df['cluster_nombre'] = indicadores_df['cluster'].map(label_map)

                # ── 6. Scatter PCA ──────────────────────────────────────────────
                fig_cluster = px.scatter(
                    indicadores_df,
                    x='pca1', y='pca2',
                    color='cluster_nombre',
                    hover_data=['nombre', 'porcentaje', 'total', 'completados'],
                    title=f'Clusters de empleados (k = {k_opt})',
                    labels={'pca1': f'PC1 ({pca.explained_variance_ratio_[0]*100:.1f}%)',
                            'pca2': f'PC2 ({pca.explained_variance_ratio_[1]*100:.1f}%)'},
                    color_discrete_sequence=px.colors.qualitative.Set2,
                    width=700, height=500,
                )
                fig_cluster.update_traces(marker=dict(size=11, line=dict(width=1, color='DarkSlateGrey')))
                fig_cluster.update_layout(legend_title_text='Cluster')
                st.plotly_chart(fig_cluster, width='stretch')

                # ── 7. Estadísticas por cluster ─────────────────────────────────
                st.markdown("**📊 Estadísticas por cluster**")
                cluster_stats = (
                    indicadores_df
                    .groupby('cluster_nombre')
                    .agg(
                        cantidad            = ('nombre',       'count'),
                        cumplimiento_prom   = ('porcentaje',   'mean'),
                        total_prom          = ('total',        'mean'),
                        completados_prom    = ('completados',  'mean'),
                        vencidos_prom       = ('vencidos',     'mean'),
                        proximos_prom       = ('proximos',     'mean'),
                    )
                    .round(1)
                    .loc[nombres]   # orden lógico bajo → alto
                )
                st.dataframe(cluster_stats, width='stretch')

                # ── 8. Análisis automático ──────────────────────────────────────
                peor  = cluster_stats.iloc[0]
                mejor = cluster_stats.iloc[-1]

                lineas_intermedias = ""
                if k_opt > 3:
                    partes = [
                        f"{row.name} ({row['cumplimiento_prom']:.1f} %)"
                        for _, row in cluster_stats.iloc[1:-1].iterrows()
                    ]
                    lineas_intermedias = f"\n- **Grupos intermedios:** {', '.join(partes)}."

                st.info(
                    f"**📌 Segmentación en {k_opt} grupos**\n\n"
                    f"- **Mayor cumplimiento:** {mejor.name} — "
                    f"{mejor['cumplimiento_prom']:.1f} % promedio, {int(mejor['cantidad'])} empleados. "
                    "Referente para mentorías y buenas prácticas.\n"
                    f"- **Menor cumplimiento:** {peor.name} — "
                    f"{peor['cumplimiento_prom']:.1f} % promedio, {int(peor['cantidad'])} empleados. "
                    "**Requiere intervención inmediata.**"
                    + lineas_intermedias +
                    "\n\n**Recomendaciones:**\n"
                    f"- Priorizar plan formativo para **{peor.name}** (seguimiento personalizado, re-entrenamientos).\n"
                    + (f"- Implementar planes de mejora progresivos para grupos intermedios.\n" if k_opt > 2 else "") +
                    f"- Potenciar a **{mejor.name}** como embajadores de buenas prácticas."
                )

            except ImportError:
                st.error("scikit-learn no encontrado. Ejecuta: python -m pip install scikit-learn")
            except Exception as e:
                st.error(f"Error en la segmentación: {e}")

            # ── 9. Interpretación de centroides ─────────────────────────────
            st.markdown("**🔍 ¿Qué define realmente a cada cluster?**")

            # Normalizar cada columna entre 0 y 1 para comparar en la misma escala
            cols_analisis = ['cumplimiento_prom', 'total_prom', 'completados_prom', 'vencidos_prom', 'proximos_prom']
            cs_norm = cluster_stats[cols_analisis].copy()
            cs_norm = (cs_norm - cs_norm.min()) / (cs_norm.max() - cs_norm.min() + 1e-9)

            etiquetas = {
                'cumplimiento_prom': '% Cumplimiento',
                'total_prom':        'Total cursos',
                'completados_prom':  'Completados',
                'vencidos_prom':     'Vencidos',
                'proximos_prom':     'Próximos a vencer',
            }

            for cluster_nombre in nombres:
                fila      = cluster_stats.loc[cluster_nombre]
                fila_norm = cs_norm.loc[cluster_nombre]

                # Feature más alto y más bajo (normalizado) para este cluster
                feat_max = fila_norm.idxmax()
                feat_min = fila_norm.idxmin()

                with st.expander(f"**{cluster_nombre}** — {int(fila['cantidad'])} empleados · {fila['cumplimiento_prom']:.1f}% cumplimiento promedio"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("Cursos totales prom.",   f"{fila['total_prom']:.1f}")
                        st.metric("Completados prom.",       f"{fila['completados_prom']:.1f}")
                        st.metric("% Cumplimiento prom.",    f"{fila['cumplimiento_prom']:.1f}%")
                    with col2:
                        st.metric("Vencidos prom.",          f"{fila['vencidos_prom']:.1f}")
                        st.metric("Próximos a vencer prom.", f"{fila['proximos_prom']:.1f}")

                    st.caption(
                        f"🔺 **Rasgo dominante:** {etiquetas[feat_max]} es el valor más alto relativo de este grupo.  \n"
                        f"🔻 **Rasgo menor:** {etiquetas[feat_min]} es el valor más bajo relativo de este grupo."
                    )
            
            # ----- 5. Tabla detallada con expansión por empleado -----
            st.subheader("📋 Detalle por Empleado (clic para expandir)")
            indicadores_df = indicadores_df.sort_values('nombre')
            tabla = indicadores_df[['nombre', 'total', 'completados', 'porcentaje', 'vigentes', 'vencidos', 'proximos', 'cluster']].copy()
            tabla['porcentaje'] = tabla['porcentaje'].round(1)
            
            for idx, row in tabla.iterrows():
                with st.expander(f"👤 {row['nombre']}  -  Cumplimiento: {row['porcentaje']}% "):
                    emp_asignaciones = asignaciones_df[asignaciones_df['empleado'] == row['nombre']]
                    if not emp_asignaciones.empty:
                        estado_counts = emp_asignaciones['estado_logico'].value_counts().reset_index()
                        estado_counts.columns = ['Estado', 'Cantidad']
                        colores = {
                            'Vigente': '#28a745',
                            'Próximo a vencer': '#ffc107',
                            'Vencido': '#dc3545',
                            'Sin vencimiento': '#6c757d',
                            'Sin ejecución': '#17a2b8'
                        }
                        fig_emp = px.bar(
                            estado_counts,
                            x='Estado',
                            y='Cantidad',
                            color='Estado',
                            color_discrete_map=colores,
                            title=f"Distribución de estados para {row['nombre']}",
                            text='Cantidad'
                        )
                        fig_emp.update_traces(textposition='outside')
                        st.plotly_chart(fig_emp, width='stretch')
                        
                        st.markdown("**Cursos asignados:**")
                        st.dataframe(
                            emp_asignaciones[['curso', 'cliente', 'fecha_asignacion', 'fecha_vencimiento', 'estado_logico']],
                            column_config={
                                'fecha_asignacion': st.column_config.DateColumn('Fecha asignación', format='DD/MM/YYYY'),
                                'fecha_vencimiento': st.column_config.DateColumn('Fecha vencimiento', format='DD/MM/YYYY'),
                                'estado_logico': 'Estado'
                            },
                            hide_index=True,
                            width='stretch'
                        )
                    else:
                        st.info("Este empleado no tiene asignaciones.")
            
        else:
            st.info("No hay datos de indicadores. Asigne cursos a empleados para visualizar.")

    # --------------------- INFORME GERENCIAL (por Curso) ---------------------
    with tabs[4]:
        st.subheader("📊 Informe Gerencial de Capacitaciones por Curso")
        st.markdown("Análisis detallado del desempeño de cada curso, tasas de finalización y distribución por cliente.")
        
        # ----- Carga de datos -----
        asignaciones = db.obtener_asignaciones()
        if asignaciones.empty:
            st.info("No hay asignaciones para generar el informe gerencial.")
        else:
            # Preparar datos
            asignaciones['fecha_asignacion'] = pd.to_datetime(asignaciones['fecha_asignacion'])
            asignaciones['fecha_vencimiento'] = pd.to_datetime(asignaciones['fecha_vencimiento'])
            hoy = pd.Timestamp(date.today())
            limite = hoy + pd.Timedelta(days=30)
            
            # ---- Estado lógico por curso ----
            def estado_logico_curso(row):
                if row['fecha_vencimiento'].year >= 9999:
                    return "Sin vencimiento"
                if row['fecha_vencimiento'] < hoy:
                    return "Vencido"
                if row['fecha_vencimiento'] <= limite:
                    return "Próximo a vencer"
                return "Vigente"
            
            asignaciones['estado_logico'] = asignaciones.apply(estado_logico_curso, axis=1)
            
            # ---- Agregar por curso ----
            # Métricas por curso
            curso_stats = asignaciones.groupby('curso').agg(
                total_asignados=('empleado', 'count'),
                completados=('estado', lambda x: (x == 'completado').sum()),
                sin_ejecucion=('estado_logico', lambda x: (x == 'Sin ejecución').sum()), 
                vencidos=('estado_logico', lambda x: (x == 'Vencido').sum()),
                vigentes=('estado_logico', lambda x: (x == 'Vigente').sum()),
                proximos=('estado_logico', lambda x: (x == 'Próximo a vencer').sum()),
                sin_vencimiento=('estado_logico', lambda x: (x == 'Sin vencimiento').sum())
            ).reset_index()
            
            # Calcular tasa de finalización
            curso_stats['tasa_finalizacion'] = (curso_stats['completados'] / curso_stats['total_asignados'] * 100).round(1)
            curso_stats['tasa_finalizacion'] = curso_stats['tasa_finalizacion'].fillna(0)
            
            # ---- Agregar cliente y duración ----
            # Obtener cliente y vigencia de cada curso desde la tabla cursos
            cursos_df = db.obtener_cursos_por_cliente()
            curso_info = cursos_df[['nombre', 'cliente', 'duracion_horas', 'vigencia_dias']].rename(columns={'nombre': 'curso'})
            curso_stats = curso_stats.merge(curso_info, on='curso', how='left')
            
            # ---- KPI globales ----
            total_cursos = len(curso_stats)
            total_asignaciones = curso_stats['total_asignados'].sum()
            total_completados = curso_stats['completados'].sum()
            tasa_global = (total_completados / total_asignaciones * 100) if total_asignaciones > 0 else 0
            
            # ---- Mejor y peor curso (considerando mínimo de asignaciones) ----
            MIN_ASIGNACIONES = 3
            cursos_con_minimo = curso_stats[curso_stats['total_asignados'] >= MIN_ASIGNACIONES]

            if not cursos_con_minimo.empty:
                tasa_max = cursos_con_minimo['tasa_finalizacion'].max()
                tasa_min = cursos_con_minimo['tasa_finalizacion'].min()

                # Todos los que empatan en tasa máxima → ordenados por asignaciones desc
                candidatos_mejor = (
                    cursos_con_minimo[cursos_con_minimo['tasa_finalizacion'] == tasa_max]
                    .sort_values('total_asignados', ascending=False)
                )

                # Todos los que empatan en tasa mínima → ordenados por asignaciones desc
                candidatos_peor = (
                    cursos_con_minimo[cursos_con_minimo['tasa_finalizacion'] == tasa_min]
                    .sort_values('total_asignados', ascending=False)
                )

                # Representante principal = el de más asignaciones
                mejor_curso     = candidatos_mejor.iloc[0]
                peor_curso      = candidatos_peor.iloc[0]

                curso_mas_exitoso   = mejor_curso['curso']
                tasa_mejor          = mejor_curso['tasa_finalizacion']
                asign_mejor         = mejor_curso['total_asignados']

                curso_menos_exitoso = peor_curso['curso']
                tasa_peor           = peor_curso['tasa_finalizacion']
                asign_peor          = peor_curso['total_asignados']

                # Listas de empate (excluyendo el representante ya mostrado)
                otros_mejor = candidatos_mejor.iloc[1:]
                otros_peor  = candidatos_peor.iloc[1:]

            else:
                curso_mas_exitoso = curso_menos_exitoso = "N/A"
                tasa_mejor = asign_mejor = tasa_peor = asign_peor = 0
                otros_mejor = otros_peor = []
            
            # ---- Mostrar KPIs en dos filas ----
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("📚 Cursos", total_cursos)
            col2.metric("📋 Asignaciones", total_asignaciones)
            col3.metric("✅ Completados", total_completados)
            col4.metric("📊 Tasa Global", f"{tasa_global:.1f}%")
            
            col5, col6 = st.columns(2)
            
            # Mostrar mejor curso con detalles
            if curso_mas_exitoso != "N/A":
                col5.metric(
                    "🏆 Mejor curso",
                    curso_mas_exitoso,
                    f"{tasa_mejor:.1f}% ({int(asign_mejor)} asignaciones)"
                )
                # Si hay más cursos con la misma tasa, mostrarlos en expander
                if len(otros_mejor) > 0:
                    with col5.expander(f"Ver todos con {tasa_mejor:.1f}% ({len(otros_mejor) + 1} cursos)"):
                        for _, row in otros_mejor.iterrows():
                            st.markdown(f"- **{row['curso']}** — {int(row['total_asignados'])} asignaciones")
            else:
                col5.metric("🏆 Mejor curso", "N/A")

            # ---- Curso con menor tasa ----
            if curso_menos_exitoso != "N/A":
                col6.metric(
                    "📉 Curso con menor tasa",
                    curso_menos_exitoso,
                    f"{tasa_peor:.1f}% ({int(asign_peor)} asignaciones)"
                )
                if len(otros_peor) > 0:
                    with col6.expander(f"Ver todos con {tasa_peor:.1f}% ({len(otros_peor) + 1} cursos)"):
                        for _, row in otros_peor.iterrows():
                            st.markdown(f"- **{row['curso']}** — {int(row['total_asignados'])} asignaciones")
            else:
                col6.metric("📉 Curso con menor tasa", "N/A")

            st.markdown("---")
            
            # ---- Filtros ----
            col_filtro1, col_filtro2 = st.columns([2, 1])
            with col_filtro1:
                cliente_options = ["Todos"] + sorted(curso_stats['cliente'].dropna().unique().tolist())
                cliente_seleccionado = st.selectbox("Filtrar por cliente", cliente_options, key="filtro_cliente_gerencial")
            with col_filtro2:
                min_tasa = st.slider("Tasa mínima de finalización (%)", 0, 100, 0, key="tasa_min_gerencial")
            
            # Aplicar filtros
            df_filtrado = curso_stats.copy()
            if cliente_seleccionado != "Todos":
                df_filtrado = df_filtrado[df_filtrado['cliente'] == cliente_seleccionado]
            df_filtrado = df_filtrado[df_filtrado['tasa_finalizacion'] >= min_tasa]
            
            if df_filtrado.empty:
                st.warning("No hay cursos que coincidan con los filtros.")
            else:
                # ---- Gráfico 1: Tasa de finalización por curso (barras horizontales) ----
                st.subheader("📊 Tasa de Finalización por Curso")
                df_ordenado = df_filtrado.sort_values('tasa_finalizacion', ascending=True)
                fig_bar = px.bar(
                    df_ordenado,
                    x='tasa_finalizacion',
                    y='curso',
                    orientation='h',
                    color='tasa_finalizacion',
                    color_continuous_scale='RdYlGn',
                    range_color=[0, 100],
                    title="",
                    labels={'tasa_finalizacion': 'Tasa de finalización (%)', 'curso': ''},
                    text='tasa_finalizacion'
                )
                fig_bar.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
                fig_bar.update_layout(height=max(400, len(df_ordenado)*25), xaxis_range=[0, 105])
                st.plotly_chart(fig_bar, width='stretch')

                # ── Análisis gráfico 1 ───────────────────────────────────────────
                if not df_ordenado.empty:
                    tasa_prom_g1    = df_ordenado['tasa_finalizacion'].mean()
                    cursos_100      = df_ordenado[df_ordenado['tasa_finalizacion'] == 100]
                    cursos_0        = df_ordenado[df_ordenado['tasa_finalizacion'] == 0]
                    cursos_criticos = df_ordenado[df_ordenado['tasa_finalizacion'] < 50]
                    curso_top       = df_ordenado.iloc[-1]
                    curso_bajo      = df_ordenado.iloc[0]

                    lineas = [
                        f"- Tasa promedio de finalización: **{tasa_prom_g1:.1f}%**.",
                        f"- **{len(cursos_100)} curso(s)** con 100% de finalización"
                        + (f": {', '.join(cursos_100['curso'].tolist())}." if len(cursos_100) <= 5 else "."),
                        f"- **{len(cursos_criticos)} curso(s)** con tasa menor al 50% → requieren intervención."
                        if len(cursos_criticos) > 0 else "- Todos los cursos superan el 50% de finalización ✅.",
                    ]
                    if len(cursos_0) > 0:
                        lineas.append(
                            f"- **{len(cursos_0)} curso(s) sin ninguna finalización**: "
                            + ", ".join(cursos_0['curso'].tolist()[:5])
                            + ("..." if len(cursos_0) > 5 else "") + "."
                        )
                    if curso_top['tasa_finalizacion'] != curso_bajo['tasa_finalizacion']:
                        lineas.append(
                            f"- Mayor brecha: **{curso_top['curso']}** ({curso_top['tasa_finalizacion']:.1f}%) "
                            f"vs **{curso_bajo['curso']}** ({curso_bajo['tasa_finalizacion']:.1f}%)."
                        )

                    st.markdown("**📌 Análisis:**  \n" + "  \n".join(lineas))

                st.markdown("---")

                # ---- Gráfico 2: Heatmap Cursos vs Clientes -------------------------
                st.subheader("🗺️ Mapa de Calor: Curso vs Cliente (Tasa de finalización)")
                pivot = df_filtrado.pivot_table(
                    index='curso',
                    columns='cliente',
                    values='tasa_finalizacion'
                    # sin fill_value: NaN donde no aplica la combinación
                )
                if not pivot.empty:
                    fig_heat = px.imshow(
                        pivot,
                        text_auto='.1f',
                        aspect='auto',
                        color_continuous_scale='RdYlGn',
                        range_color=[0, 100],
                        labels=dict(x='Cliente', y='Curso', color='Tasa %')
                    )
                    fig_heat.update_layout(height=max(400, len(pivot)*25))
                    st.plotly_chart(fig_heat, width='stretch')

                    # ── Análisis gráfico 2 ───────────────────────────────────────
                    # Solo columnas/filas con al menos un dato real
                    tasa_por_cliente = pivot.mean(axis=0).dropna().sort_values(ascending=False)
                    tasa_por_curso   = pivot.mean(axis=1).dropna().sort_values(ascending=False)

                    # Celdas con dato real (no NaN) y tasa == 0
                    celdas_cero = [
                        (c, cl)
                        for c in pivot.index
                        for cl in pivot.columns
                        if pd.notna(pivot.loc[c, cl]) and pivot.loc[c, cl] == 0
                    ]

                    lineas_h = []

                    if not tasa_por_cliente.empty:
                        mejor_cli = tasa_por_cliente.index[0]
                        peor_cli  = tasa_por_cliente.index[-1]
                        lineas_h += [
                            f"- Cliente con mayor cumplimiento promedio: **{mejor_cli}** ({tasa_por_cliente[mejor_cli]:.1f}%).",
                            f"- Cliente con menor cumplimiento promedio: **{peor_cli}** ({tasa_por_cliente[peor_cli]:.1f}%) → priorizar seguimiento.",
                        ]

                    if not tasa_por_curso.empty:
                        mejor_cur = tasa_por_curso.index[0]
                        peor_cur  = tasa_por_curso.index[-1]
                        lineas_h += [
                            f"- Curso con mejor desempeño transversal: **{mejor_cur}** ({tasa_por_curso[mejor_cur]:.1f}% promedio entre clientes).",
                            f"- Curso con peor desempeño transversal: **{peor_cur}** ({tasa_por_curso[peor_cur]:.1f}% promedio entre clientes).",
                        ]

                    if celdas_cero:
                        ejemplos = ", ".join([f"{c}/{cl}" for c, cl in celdas_cero[:4]])
                        lineas_h.append(
                            f"- **{len(celdas_cero)} combinación(es) curso/cliente con 0% real**: {ejemplos}"
                            + ("..." if len(celdas_cero) > 4 else "") + "."
                        )
                    else:
                        lineas_h.append("- No hay combinaciones curso/cliente con 0% de cumplimiento real ✅.")

                    if lineas_h:
                        st.markdown("**📌 Análisis:**  \n" + "  \n".join(lineas_h))

                else:
                    st.info("No hay suficientes datos para el heatmap.")

                st.markdown("---")

                # ---- Gráfico 3: Dispersión Vigencia vs Tasa de finalización --------
                st.subheader("⏱️ Relación Vigencia del Curso vs Tasa de Finalización")
                if df_filtrado['vigencia_dias'].notna().any():
                    df_scatter = df_filtrado[df_filtrado['vigencia_dias'].notna()].copy()

                    fig_scatter = px.scatter(
                        df_scatter,
                        x='vigencia_dias',
                        y='tasa_finalizacion',
                        size='total_asignados',
                        color='cliente',
                        hover_name='curso',
                        title="",
                        labels={'vigencia_dias': 'Vigencia (días)', 'tasa_finalizacion': 'Tasa de finalización (%)'},
                        size_max=20,
                    )
                    fig_scatter.update_layout(height=500)
                    st.plotly_chart(fig_scatter, width='stretch')

                    # ── Análisis gráfico 3 ───────────────────────────────────────
                    corr = df_scatter['vigencia_dias'].corr(df_scatter['tasa_finalizacion'])

                    if   corr >  0.4:  texto_corr = f"correlación positiva moderada-alta ({corr:.2f}) — cursos con mayor vigencia tienden a completarse más."
                    elif corr < -0.4:  texto_corr = f"correlación negativa moderada-alta ({corr:.2f}) — cursos con mayor vigencia tienden a completarse menos."
                    else:               texto_corr = f"correlación débil ({corr:.2f}) — la vigencia no explica significativamente la tasa de finalización."

                    vig_media  = df_scatter['vigencia_dias'].median()
                    tasa_media = df_scatter['tasa_finalizacion'].median()

                    alta_vig_alta_tasa = df_scatter[(df_scatter['vigencia_dias'] >= vig_media) & (df_scatter['tasa_finalizacion'] >= tasa_media)]
                    baja_vig_alta_tasa = df_scatter[(df_scatter['vigencia_dias'] <  vig_media) & (df_scatter['tasa_finalizacion'] >= tasa_media)]
                    alta_vig_baja_tasa = df_scatter[(df_scatter['vigencia_dias'] >= vig_media) & (df_scatter['tasa_finalizacion'] <  tasa_media)]
                    baja_vig_baja_tasa = df_scatter[(df_scatter['vigencia_dias'] <  vig_media) & (df_scatter['tasa_finalizacion'] <  tasa_media)]

                    lineas_s = [
                        f"- Existe una {texto_corr}",
                        f"- Vigencia mediana: **{vig_media:.0f} días** | Tasa mediana: **{tasa_media:.1f}%**.",
                        f"- **Cuadrante ideal** (alta vigencia + alta tasa): {len(alta_vig_alta_tasa)} curso(s).",
                        f"- **Cursos eficientes** (baja vigencia + alta tasa): {len(baja_vig_alta_tasa)} curso(s) → buena relación tiempo/resultado.",
                        f"- **Cursos en riesgo** (alta vigencia + baja tasa): {len(alta_vig_baja_tasa)} curso(s) → larga duración sin resultados, revisar diseño.",
                        f"- **Cursos críticos** (baja vigencia + baja tasa): {len(baja_vig_baja_tasa)} curso(s) → prioridad de intervención.",
                    ]

                    sin_vigencia = df_filtrado[df_filtrado['vigencia_dias'].isna()]
                    if not sin_vigencia.empty:
                        lineas_s.append(
                            f"- ℹ️ {len(sin_vigencia)} curso(s) sin vigencia registrada no aparecen en el gráfico."
                        )

                    st.markdown("**📌 Análisis:**  \n" + "  \n".join(lineas_s))

                else:
                    st.info("📌 No hay datos de vigencia para mostrar la relación con la tasa de finalización.")

                
                # ---- Tabla interactiva detallada ----
                st.subheader("📋 Tabla Detallada de Cursos")
                # Ordenar por tasa descendente
                df_tabla = df_filtrado.sort_values('tasa_finalizacion', ascending=False)
                # Renombrar columnas (excluyendo Vigentes, Sin vencimiento, Duración)
                df_tabla = df_tabla.rename(columns={
                    'curso': 'Curso',
                    'cliente': 'Cliente',
                    'total_asignados': 'Asignados',
                    'completados': 'Completados',
                    'sin_ejecucion': 'Sin ejecución',
                    'tasa_finalizacion': 'Tasa %',
                    'vencidos': 'Vencidos',
                    'proximos': 'Próximos (≤30 días)',
                    'vigencia_dias': 'Vigencia (días)'
                })
                # Formatear la columna Tasa %
                df_tabla['Tasa %'] = df_tabla['Tasa %'].map('{:.1f}%'.format)
                
                st.dataframe(
                    df_tabla[['Curso', 'Cliente', 'Asignados', 'Completados', 'Sin ejecución', 'Tasa %', 'Vencidos', 'Próximos (≤30 días)', 'Vigencia (días)']],
                    column_config={
                        "Tasa %": st.column_config.Column("Tasa %", width="small"),
                        "Asignados": st.column_config.Column("Asignados", width="small"),
                        "Completados": st.column_config.Column("Completados", width="small"),
                        "Sin ejecución": st.column_config.Column("Sin ejecución", width="small"),   
                        "Vencidos": st.column_config.Column("Vencidos", width="small"),
                        "Próximos (≤30 días)": st.column_config.Column("Próximos", width="small"),
                        "Vigencia (días)": st.column_config.Column("Vigencia (días)", width="small")
                    },
                    hide_index=True,
                    width='stretch'
                )

                
            # ---- Exportar a Excel profesional ----
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            def generar_excel_gerencial(df_filtrado, total_cursos, total_asignaciones, 
                                        total_completados, tasa_global, curso_mas_exitoso,
                                        curso_menos_exitoso):
                output = io.BytesIO()
                wb = Workbook()
                
                # ----- HOJA 1: RESUMEN EJECUTIVO -----
                ws_resumen = wb.active
                ws_resumen.title = "Resumen Ejecutivo"
                
                # Estilos
                header_font = Font(bold=True, size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Título
                ws_resumen['A1'] = "INFORME GERENCIAL DE CAPACITACIONES"
                ws_resumen['A1'].font = Font(bold=True, size=16)
                ws_resumen.merge_cells('A1:D1')
                
                # Fecha
                ws_resumen['A2'] = f"Fecha de generación: {date.today().strftime('%d/%m/%Y')}"
                ws_resumen.merge_cells('A2:D2')
                
                # KPIs
                kpis = [
                    ("Métrica", "Valor"),
                    ("Total cursos", total_cursos),
                    ("Total asignaciones", total_asignaciones),
                    ("Cursos completados", total_completados),
                    ("Tasa de finalización global", f"{tasa_global:.1f}%"),
                    ("Mejor curso", curso_mas_exitoso),
                    ("Curso con menor tasa", curso_menos_exitoso)
                ]
                
                row = 4
                for item in kpis:
                    ws_resumen.cell(row=row, column=1, value=item[0]).font = header_font
                    ws_resumen.cell(row=row, column=2, value=item[1])
                    row += 1
                
                ws_resumen.column_dimensions['A'].width = 30
                ws_resumen.column_dimensions['B'].width = 20
                
                # ----- HOJA 2: DETALLE POR CURSO -----
                ws_cursos = wb.create_sheet("Detalle por Curso")
                # Renombrar columnas para la tabla
                df_tabla = df_filtrado.rename(columns={
                    'curso': 'Curso',
                    'cliente': 'Cliente',
                    'total_asignados': 'Asignados',
                    'completados': 'Completados',
                    'tasa_finalizacion': 'Tasa %',
                    'vigentes': 'Vigentes',
                    'vencidos': 'Vencidos',
                    'proximos': 'Próximos',
                    'sin_vencimiento': 'Sin vencimiento',
                    'duracion_horas': 'Duración (h)',
                    'vigencia_dias': 'Vigencia (días)'
                })
                # Formatear la columna Tasa % como string con '%'
                df_tabla['Tasa %'] = df_tabla['Tasa %'].map('{:.1f}%'.format)
                
                # Encabezados
                headers = list(df_tabla.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_cursos.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                # Datos
                for r_idx, row_data in enumerate(df_tabla.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = ws_cursos.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center" if isinstance(value, (int, float)) else "left")
                
                # Ajustar anchos
                for col in ws_cursos.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws_cursos.column_dimensions[column].width = adjusted_width
                
                # ----- HOJA 3: ANÁLISIS POR CLIENTE -----
                ws_cliente = wb.create_sheet("Por Cliente")
                
                # Agrupar por cliente (usando columna original 'cliente')
                cliente_resumen = df_filtrado.groupby('cliente').agg({
                    'total_asignados': 'sum',
                    'completados': 'sum',
                    'tasa_finalizacion': 'mean'
                }).round(1).reset_index()
                cliente_resumen.columns = ['Cliente', 'Asignados', 'Completados', 'Tasa media %']
                # Formatear tasa media
                cliente_resumen['Tasa media %'] = cliente_resumen['Tasa media %'].map('{:.1f}%'.format)
                
                # Escribir encabezados
                for col_idx, header in enumerate(cliente_resumen.columns, 1):
                    cell = ws_cliente.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                # Datos
                for r_idx, row_data in enumerate(cliente_resumen.itertuples(index=False), 2):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = ws_cliente.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")
                
                ws_cliente.column_dimensions['A'].width = 25
                ws_cliente.column_dimensions['B'].width = 15
                ws_cliente.column_dimensions['C'].width = 15
                ws_cliente.column_dimensions['D'].width = 18
                
                wb.save(output)
                output.seek(0)
                return output
            
            # Generar Excel
            excel_data = generar_excel_gerencial(
                df_filtrado,
                total_cursos,
                total_asignaciones,
                total_completados,
                tasa_global,
                curso_mas_exitoso,
                curso_menos_exitoso
            )
            
            st.download_button(
                label="📥 Descargar informe en Excel",
                data=excel_data,
                file_name=f'informe_gerencial_{date.today().strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='download_excel_gerencial'
            )


            st.markdown("---")
            
            # ---- ANÁLISIS POR CLIENTE (APC, Ecopetrol, Otros) ----
            st.subheader("📊 Análisis de Cumplimiento por Grupo de Cliente")
            
            # Calcular la tasa global para usarla como referencia de variación
            tasa_global_valor = (total_completados / total_asignaciones * 100) if total_asignaciones > 0 else 0
            
            # Función para formatear variación con signo y color
            def formatear_variacion(tasa, referencia=tasa_global_valor):
                diff = tasa - referencia
                if abs(diff) < 0.1:
                    return "0.0%"
                return f"{'+' if diff > 0 else ''}{diff:.1f}%"
            
            # ---- 1. Clasificar cursos por cliente ----
            # Definir grupos de clientes
            clientes_apc = ['APC']
            clientes_ecopetrol = ['ECOPETROL'] + [cli for cli in df_filtrado['cliente'].unique() if cli.startswith('ECP -')]
            
            # Filtrar DataFrames
            df_apc = df_filtrado[df_filtrado['cliente'].isin(clientes_apc)].copy()
            df_ecopetrol = df_filtrado[df_filtrado['cliente'].isin(clientes_ecopetrol)].copy()
            df_otros = df_filtrado[~df_filtrado['cliente'].isin(clientes_apc + clientes_ecopetrol)].copy()
            
            # ---- 2. Mostrar tablas por grupo ----
            # Función para crear tabla de cursos
            def mostrar_tabla_cursos(df, titulo):
                if df.empty:
                    st.info(f"No hay cursos en el grupo **{titulo}**.")
                    return
                df_tabla = df.sort_values('tasa_finalizacion', ascending=False)
                df_tabla['Tasa %'] = df_tabla['tasa_finalizacion'].map('{:.1f}%'.format)
                df_tabla['Variación'] = df_tabla['tasa_finalizacion'].apply(
                    lambda x: formatear_variacion(x, tasa_global_valor)
                )
                st.markdown(f"### 📌 {titulo}")
                st.dataframe(
                    df_tabla[['curso', 'Tasa %', 'Variación']],
                    column_config={
                        "curso": "Capacitación",
                        "Tasa %": st.column_config.Column("Cumplimiento", width="small"),
                        "Variación": st.column_config.Column("Variación", width="small")
                    },
                    hide_index=True,
                    width='stretch'
                )
            
            # Mostrar cada grupo
            mostrar_tabla_cursos(df_apc, "Capacitación General (APC)")
            mostrar_tabla_cursos(df_ecopetrol, "Ecopetrol")
            mostrar_tabla_cursos(df_otros, "Otros Clientes")
            
            # ---- 3. Tabla de Clientes ----
            st.markdown("### 🏢 Cumplimiento por Cliente")
            cliente_stats = df_filtrado.groupby('cliente').agg(
                asignados=('total_asignados', 'sum'),
                completados=('completados', 'sum'),
                tasa=('tasa_finalizacion', lambda x: (x.sum() / len(x)) if len(x) > 0 else 0)
            ).reset_index()
            cliente_stats['tasa'] = cliente_stats['tasa'].round(1)
            cliente_stats['Tasa %'] = cliente_stats['tasa'].map('{:.1f}%'.format)
            cliente_stats['Variación'] = cliente_stats['tasa'].apply(
                lambda x: formatear_variacion(x, tasa_global_valor)
            )
            st.dataframe(
                cliente_stats[['cliente', 'Tasa %', 'Variación']],
                column_config={
                    "cliente": "Cliente",
                    "Tasa %": st.column_config.Column("Cumplimiento", width="small"),
                    "Variación": st.column_config.Column("Variación", width="small")
                },
                hide_index=True,
                width='stretch'
            )
            
            # ---- 4. Gerencias Ecopetrol ----
            st.markdown("### 🏢 Gerencias Ecopetrol")
            df_ecopetrol_gerencias = df_filtrado[df_filtrado['cliente'].str.startswith('ECP -', na=False)].copy()
            if not df_ecopetrol_gerencias.empty:
                gerencia_stats = df_ecopetrol_gerencias.groupby('cliente').agg(
                    asignados=('total_asignados', 'sum'),
                    completados=('completados', 'sum'),
                    tasa=('tasa_finalizacion', lambda x: (x.sum() / len(x)) if len(x) > 0 else 0)
                ).reset_index()
                gerencia_stats['tasa'] = gerencia_stats['tasa'].round(1)
                gerencia_stats['Tasa %'] = gerencia_stats['tasa'].map('{:.1f}%'.format)
                gerencia_stats['Variación'] = gerencia_stats['tasa'].apply(
                    lambda x: formatear_variacion(x, tasa_global_valor)
                )
                # Agregar fila "GENERAL" con el promedio de todas las gerencias
                promedio_gerencias = gerencia_stats['tasa'].mean()
                if not gerencia_stats.empty:
                    general_row = pd.DataFrame({
                        'cliente': ['GENERAL'],
                        'Tasa %': [f"{promedio_gerencias:.1f}%"],
                        'Variación': [formatear_variacion(promedio_gerencias, tasa_global_valor)]
                    })
                    gerencia_stats_display = gerencia_stats[['cliente', 'Tasa %', 'Variación']]
                    gerencia_stats_display = pd.concat([gerencia_stats_display, general_row], ignore_index=True)
                else:
                    gerencia_stats_display = gerencia_stats[['cliente', 'Tasa %', 'Variación']]
                
                st.dataframe(
                    gerencia_stats_display,
                    column_config={
                        "cliente": "Gerencia",
                        "Tasa %": st.column_config.Column("Cumplimiento", width="small"),
                        "Variación": st.column_config.Column("Variación", width="small")
                    },
                    hide_index=True,
                    width='stretch'
                )
            else:
                st.info("No hay gerencias de Ecopetrol en el filtro actual.")
            
            st.markdown("---")
            


        with tabs[5]:
            st.subheader("📅 Proyección del Cronograma - Cursos por Vencer")
            st.markdown("Visualiza los cursos que vencen en el período seleccionado, organizados por empleado.")

            asignaciones = db.obtener_asignaciones()
            if asignaciones.empty:
                st.info("No hay asignaciones para mostrar.")
            else:
                asignaciones['fecha_vencimiento'] = pd.to_datetime(asignaciones['fecha_vencimiento'])
                hoy = pd.Timestamp(date.today())

                # ── Filtros de período ──────────────────────────────────────────
                col_f1, col_f2, col_f3 = st.columns([1, 1, 2])

                with col_f1:
                    tipo_filtro = st.selectbox(
                        "Período",
                        ["Día", "Mes", "Semestre", "Año"],
                        index=1,
                        key="cal_tipo_filtro"
                    )

                with col_f2:
                    # Asegurar que el año actual y algunos futuros estén en la lista
                    anios_disponibles = sorted(
                        asignaciones['fecha_vencimiento']
                        .dt.year
                        .unique()
                        .tolist()
                    )
                    # Agregar años futuros manualmente (hasta 2030 por ejemplo)
                    años_futuros = [hoy.year + i for i in range(0, 6)]  # 0 a 5 años adelante
                    for y in años_futuros:
                        if y not in anios_disponibles:
                            anios_disponibles.append(y)
                    anios_disponibles.sort()

                    anio_sel = st.selectbox(
                        "Año",
                        anios_disponibles,
                        index=anios_disponibles.index(hoy.year),
                        key="cal_anio"
                    )

                with col_f3:
                    if tipo_filtro == "Día":
                        dia_sel = st.date_input(
                            "Fecha",
                            value=hoy.date(),
                            key="cal_dia"
                        )
                        fecha_ini = pd.Timestamp(dia_sel)
                        fecha_fin = fecha_ini
                        mes_sel = fecha_ini.month
                        anio_sel = fecha_ini.year

                    elif tipo_filtro == "Mes":
                        MESES_ES = {
                            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
                        }
                        mes_sel = st.selectbox(
                            "Mes",
                            list(MESES_ES.keys()),
                            format_func=lambda m: MESES_ES[m],
                            index=hoy.month - 1,
                            key="cal_mes"
                        )
                        fecha_ini = pd.Timestamp(year=anio_sel, month=mes_sel, day=1)
                        fecha_fin = fecha_ini + pd.offsets.MonthEnd(0)

                    elif tipo_filtro == "Semestre":
                        sem_sel = st.radio(
                            "Semestre",
                            ["1er semestre (Ene–Jun)", "2do semestre (Jul–Dic)"],
                            horizontal=True,
                            key="cal_semestre"
                        )
                        if "1er" in sem_sel:
                            fecha_ini = pd.Timestamp(year=anio_sel, month=1, day=1)
                            fecha_fin = pd.Timestamp(year=anio_sel, month=6, day=30)
                            mes_sel = 1
                        else:
                            fecha_ini = pd.Timestamp(year=anio_sel, month=7, day=1)
                            fecha_fin = pd.Timestamp(year=anio_sel, month=12, day=31)
                            mes_sel = 7

                    else:  # Año
                        st.write("")  # spacer
                        fecha_ini = pd.Timestamp(year=anio_sel, month=1, day=1)
                        fecha_fin = pd.Timestamp(year=anio_sel, month=12, day=31)
                        mes_sel = 1

                # ── FILTRO PRINCIPAL: TODOS los cursos con vencimiento en el período ──
                mask = (
                    (asignaciones['fecha_vencimiento'] >= fecha_ini) &
                    (asignaciones['fecha_vencimiento'] <= fecha_fin) &
                    (asignaciones['fecha_vencimiento'].dt.year < 9999)   # excluye sin vencimiento
                )
                por_vencer = asignaciones[mask].copy()

                # ── Mensaje informativo si no hay datos ──
                if por_vencer.empty:
                    # Buscar el próximo mes con vencimientos (igual que antes)
                    futuros = asignaciones[
                        (asignaciones['fecha_vencimiento'] >= hoy) &
                        (asignaciones['estado'] != 'completado') &
                        (asignaciones['fecha_vencimiento'].dt.year < 9999)
                    ]
                    if not futuros.empty:
                        futuros['anio_mes'] = futuros['fecha_vencimiento'].dt.to_period('M')
                        meses_con_datos = sorted(futuros['anio_mes'].unique(), key=lambda x: (x.year, x.month))
                        mes_seleccionado = pd.Period(year=anio_sel, month=mes_sel, freq='M')
                        prox_mes = None
                        for m in meses_con_datos:
                            if m >= mes_seleccionado:
                                prox_mes = m
                                break
                        if prox_mes:
                            meses_es = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                                        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                                        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
                            nombre_mes_sel = meses_es[mes_sel]
                            nombre_mes_prox = meses_es[prox_mes.month]
                            st.info(f"ℹ️ No hay cursos con vencimiento en **{nombre_mes_sel} de {anio_sel}**. El próximo mes con vencimientos es **{nombre_mes_prox} de {prox_mes.year}**.")
                        else:
                            st.info("✅ No hay cursos futuros en el sistema.")
                    else:
                        st.info("✅ No hay cursos futuros en el sistema.")

                # ── Categorización para el calendario ────────────────────────
                def categoria_calendario(row):
                    diff = (row['fecha_vencimiento'] - hoy).days
                    if row['estado'] == 'completado':
                        return None  # excluir completados
                    if diff < 0:
                        return 'vencido'
                    elif diff <= 3:
                        return 'urgente'
                    elif diff <= 30:
                        return 'proximo'
                    else:
                        return 'normal'

                # ── Preparar eventos JSON ────────────────────────────────────
                if por_vencer.empty:
                    eventos_json = "[]"
                else:
                    por_vencer['categoria'] = por_vencer.apply(categoria_calendario, axis=1)
                    por_vencer['dias_restantes_calc'] = (por_vencer['fecha_vencimiento'] - hoy).dt.days
                    # Solo incluimos los que no son 'normal' porque la tabla de detalle los muestra, pero para las tarjetas solo queremos vencidos y próximas (urgente+proximo)
                    # Pero el calendario muestra todos los del período (incluyendo normales) para tener visión completa.
                    eventos_json = json.dumps([
                        {
                            "empleado": str(row['empleado']),
                            "curso":    str(row['curso']),
                            "cliente":  str(row['cliente']),
                            "fecha":    row['fecha_vencimiento'].strftime("%Y-%m-%dT00:00:00"),
                            "estado":   str(row['estado']),
                            "categoria": row['categoria'],
                            "dias":     int(row['dias_restantes_calc'])
                        }
                        for _, row in por_vencer.iterrows()
                    ], ensure_ascii=False)

                # ── Estadísticas del período (contando TODOS los cursos, sin excluir completados) ──
                if por_vencer.empty:
                    stats_periodo = {'vencidos': 0, 'proximas': 0}
                else:
                    # Calcular días restantes para todos (sin usar categoría)
                    por_vencer['dias_restantes'] = (por_vencer['fecha_vencimiento'] - hoy).dt.days
                    stats_periodo = {
                        'vencidos': len(por_vencer[por_vencer['dias_restantes'] < 0]),
                        'proximas': len(por_vencer[(por_vencer['dias_restantes'] >= 0) & (por_vencer['dias_restantes'] <= 30)])
                    }

                # Variables para el JavaScript
                cal_year = anio_sel
                cal_month = mes_sel - 1
                _vencidos = stats_periodo['vencidos']
                _proximas = stats_periodo['proximas']

                # ── HTML del calendario con solo dos tarjetas ──────────────────
                calendario_html = f"""
                <!DOCTYPE html>
                <html lang="es">
                <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <script>
                const stats_vencidos = {_vencidos};
                const stats_proximas = {_proximas};
                </script>
                <style>
                * {{ box-sizing: border-box; margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }}
                body {{ background: transparent; padding: 0; }}

                .stats-row {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; margin-bottom: 14px; }}
                .stat {{ background: #f7f7f8; border: 0.5px solid #e0e0e0; border-radius: 10px; padding: 10px 14px; }}
                .stat-label {{ font-size: 11px; color: #888; margin-bottom: 3px; }}
                .stat-val {{ font-size: 22px; font-weight: 600; color: #1a1a1a; }}
                .stat-val.rojo {{ color: #a32d2d; }}
                .stat-val.ambar {{ color: #854f0b; }}

                /* Resto de estilos (cal-header, week-header, cal-grid, etc.) se mantienen igual */
                .cal-header {{ display: flex; align-items: center; justify-content: space-between; margin-bottom: 10px; flex-wrap: wrap; gap: 8px; }}
                .cal-period {{ font-size: 16px; font-weight: 600; color: #1a1a1a; }}
                .view-btns {{ display: flex; gap: 4px; }}
                .vbtn {{ background: #f0f0f2; border: 0.5px solid #ddd; border-radius: 8px; padding: 5px 14px; font-size: 12px; font-weight: 500; color: #555; cursor: pointer; }}
                .vbtn.active {{ background: #fff; border-color: #378ADD; color: #185fa5; font-weight: 600; }}

                .week-header {{ display: grid; grid-template-columns: repeat(7, 1fr); text-align: center; margin-bottom: 3px; }}
                .week-header span {{ font-size: 11px; font-weight: 600; color: #aaa; padding: 4px 0; text-transform: uppercase; letter-spacing: 0.05em; }}

                .cal-grid {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 3px; }}
                .cal-day {{
                    min-height: 82px; border: 0.5px solid #ebebeb; border-radius: 9px;
                    padding: 6px 5px 4px; background: #fff; position: relative; vertical-align: top;
                }}
                .cal-day.otro-mes {{ background: #fafafa; opacity: 0.5; }}
                .cal-day.hoy {{ border-color: #378ADD; border-width: 2px; }}
                .day-num {{
                    font-size: 12px; font-weight: 600; color: #1a1a1a;
                    width: 22px; height: 22px; display: flex; align-items: center;
                    justify-content: center; border-radius: 50%; margin-bottom: 2px;
                }}
                .hoy .day-num {{ background: #378ADD; color: #fff; }}
                .ev-pill {{
                    font-size: 10px; font-weight: 500; border-radius: 4px;
                    padding: 2px 4px; margin-top: 2px; white-space: nowrap;
                    overflow: hidden; text-overflow: ellipsis; cursor: pointer;
                    border-left: 3px solid; line-height: 1.45;
                }}
                .ev-pill.vencido  {{ background: #fee2e2; color: #991b1b; border-left-color: #b91c1c; }}
                .ev-pill.urgente  {{ background: #fcebeb; color: #a32d2d; border-left-color: #e24b4a; }}
                .ev-pill.proximo  {{ background: #faeeda; color: #854f0b; border-left-color: #ef9f27; }}
                .ev-pill.normal   {{ background: #faece7; color: #993c1d; border-left-color: #d85a30; }}
                .more-tag {{ font-size: 10px; color: #999; margin-top: 2px; cursor: pointer; }}

                .list-view {{ display: none; }}
                .list-date-header {{ font-size: 12px; font-weight: 600; color: #888; padding: 10px 0 4px; text-transform: capitalize; border-bottom: 0.5px solid #eee; margin-bottom: 2px; }}
                .ev-row {{ display: flex; align-items: flex-start; gap: 10px; padding: 8px 0; border-bottom: 0.5px solid #f0f0f0; }}
                .ev-dot {{ width: 8px; height: 8px; border-radius: 50%; margin-top: 4px; flex-shrink: 0; }}
                .ev-info {{ flex: 1; min-width: 0; }}
                .ev-name {{ font-size: 13px; font-weight: 600; color: #1a1a1a; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
                .ev-meta {{ font-size: 11px; color: #888; margin-top: 1px; }}
                .ev-badge {{ font-size: 11px; padding: 2px 9px; border-radius: 20px; font-weight: 600; flex-shrink: 0; }}
                .badge-vencido  {{ background: #fee2e2; color: #991b1b; }}
                .badge-urgente  {{ background: #fcebeb; color: #a32d2d; }}
                .badge-proximo  {{ background: #faeeda; color: #854f0b; }}
                .badge-normal   {{ background: #faece7; color: #993c1d; }}

                .detail-panel {{
                    margin-top: 12px; border: 0.5px solid #e0e0e0; border-radius: 12px;
                    padding: 14px 16px; background: #fff; display: none;
                }}
                .detail-title {{ font-size: 13px; font-weight: 600; color: #888; margin-bottom: 8px; }}
                .detail-content-row {{ display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; flex-wrap: wrap; }}
                .detail-main .d-emp {{ font-size: 15px; font-weight: 600; color: #1a1a1a; }}
                .detail-main .d-cur {{ font-size: 13px; color: #555; margin-top: 3px; }}
                .detail-main .d-cli {{ font-size: 12px; color: #888; margin-top: 2px; }}
                .detail-right {{ text-align: right; }}
                .detail-right .d-badge {{ font-size: 13px; padding: 4px 14px; border-radius: 20px; font-weight: 600; display: inline-block; }}
                .detail-right .d-date {{ font-size: 12px; color: #888; margin-top: 6px; }}
                .close-btn {{ float: right; background: none; border: none; font-size: 18px; color: #aaa; cursor: pointer; margin-top: -2px; }}

                .anual-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 14px; }}
                .mini-mes {{ background: #fff; border: 0.5px solid #ebebeb; border-radius: 10px; padding: 10px; }}
                .mini-mes-title {{ font-size: 12px; font-weight: 600; color: #555; margin-bottom: 6px; text-align: center; }}
                .mini-week-hdr {{ display: grid; grid-template-columns: repeat(7, 1fr); text-align: center; margin-bottom: 2px; }}
                .mini-week-hdr span {{ font-size: 9px; color: #bbb; font-weight: 600; }}
                .mini-grid {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 1px; }}
                .mini-day {{ height: 20px; display: flex; align-items: center; justify-content: center; font-size: 9px; color: #888; border-radius: 4px; position: relative; cursor: default; }}
                .mini-day.otro-mes-m {{ color: #ddd; }}
                .mini-day.hoy-m {{ background: #378ADD; color: #fff; font-weight: 700; border-radius: 50%; }}
                .mini-day.con-eventos {{ background: #fee; border-radius: 50%; color: #c0392b; font-weight: 700; cursor: pointer; }}
                .mini-day.con-eventos:hover {{ background: #fcc; }}

                .empty-message {{
                    text-align: center;
                    color: #aaa;
                    font-size: 14px;
                    padding: 2rem 0;
                    grid-column: 1 / -1;
                }}
                </style>
                </head>
                <body>

                <div id="statsRow" class="stats-row"></div>

                <div class="cal-header">
                <span class="cal-period" id="calPeriod"></span>
                <div class="view-btns">
                    <button class="vbtn active" id="btnMes"   onclick="setView('mes')">Mes</button>
                    <button class="vbtn"        id="btnLista" onclick="setView('lista')">Lista</button>
                    <button class="vbtn"        id="btnAnual" onclick="setView('anual')">Año</button>
                </div>
                </div>

                <!-- Vista mes -->
                <div id="mesView">
                <div class="week-header">
                    <span>Dom</span><span>Lun</span><span>Mar</span><span>Mié</span><span>Jue</span><span>Vie</span><span>Sáb</span>
                </div>
                <div class="cal-grid" id="calGrid"></div>
                </div>

                <!-- Vista lista -->
                <div class="list-view" id="listaView"></div>

                <!-- Vista anual -->
                <div style="display:none" id="anualView">
                <div class="anual-grid" id="anualGrid"></div>
                </div>

                <!-- Panel detalle -->
                <div class="detail-panel" id="detailPanel">
                <button class="close-btn" onclick="document.getElementById('detailPanel').style.display='none'">&#x2715;</button>
                <div class="detail-title" id="detailTitle"></div>
                <div id="detailContent"></div>
                </div>

                <script>
                const calYear = {cal_year};
                const calMonth = {cal_month};
                const eventosJson = {eventos_json};

                const todayDate = new Date();
                todayDate.setHours(0,0,0,0);

                const eventos = eventosJson.map(e => ({{
                ...e,
                fecha: new Date(e.fecha)
                }}));

                const MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
                const SEMANAS_CORTO = ['D','L','M','X','J','V','S'];

                let currentView = 'mes';
                const viewMonth = calMonth;
                const viewYear  = calYear;

                function diffDias(fecha) {{
                    return Math.ceil((fecha - todayDate) / 864e5);
                }}

                function badgeText(ev) {{
                    const d = ev.dias;
                    if (ev.categoria === 'vencido') return `Vencido hace ${{Math.abs(d)}}d`;
                    if (ev.categoria === 'urgente') return d === 0 ? 'Hoy' : `En ${{d}}d`;
                    if (ev.categoria === 'proximo') return `En ${{d}}d`;
                    return `En ${{d}}d`;
                }}

                function dotColor(cat) {{
                    if (cat === 'vencido') return '#b91c1c';
                    if (cat === 'urgente') return '#e24b4a';
                    if (cat === 'proximo') return '#ef9f27';
                    return '#d85a30';
                }}

                function evsByDay(y, m, d) {{
                    return eventos.filter(e => e.fecha.getFullYear()===y && e.fecha.getMonth()===m && e.fecha.getDate()===d);
                }}

                function renderStats() {{
                    const v = stats_vencidos;
                    const p = stats_proximas;
                    document.getElementById('statsRow').innerHTML = `
                        <div class="stat"><div class="stat-label">❌ Vencidos</div><div class="stat-val rojo">${{v}}</div></div>
                        <div class="stat"><div class="stat-label">⚠ Próximos a vencer (≤30d)</div><div class="stat-val ambar">${{p}}</div></div>
                    `;
                }}

                function renderMes() {{
                    document.getElementById('calPeriod').textContent = `${{MESES[viewMonth]}} ${{viewYear}}`;
                    const firstDay   = new Date(viewYear, viewMonth, 1).getDay();
                    const daysInMonth= new Date(viewYear, viewMonth+1, 0).getDate();
                    const daysInPrev = new Date(viewYear, viewMonth, 0).getDate();
                    const total      = Math.ceil((firstDay + daysInMonth) / 7) * 7;
                    const grid       = document.getElementById('calGrid');
                    grid.innerHTML   = '';

                    if (eventos.length === 0) {{
                        for (let i = 0; i < total; i++) {{
                            let dayN, mo, yr, otro = false;
                            if (i < firstDay) {{
                                dayN = daysInPrev - firstDay + i + 1; mo = viewMonth-1; yr = viewYear; otro = true;
                            }} else if (i >= firstDay + daysInMonth) {{
                                dayN = i - firstDay - daysInMonth + 1; mo = viewMonth+1; yr = viewYear; otro = true;
                            }} else {{
                                dayN = i - firstDay + 1; mo = viewMonth; yr = viewYear;
                            }}
                            if (mo < 0)  {{ mo = 11; yr--; }}
                            if (mo > 11) {{ mo = 0;  yr++; }}
                            const cell = new Date(yr, mo, dayN);
                            cell.setHours(0,0,0,0);
                            const isHoy = cell.getTime() === todayDate.getTime();
                            const div = document.createElement('div');
                            div.className = 'cal-day' + (otro ? ' otro-mes' : '') + (isHoy ? ' hoy' : '');
                            div.innerHTML = `<div class="day-num">${{dayN}}</div>`;
                            grid.appendChild(div);
                        }}
                        const msgDiv = document.createElement('div');
                        msgDiv.className = 'empty-message';
                        msgDiv.textContent = 'No hay cursos en este período';
                        grid.appendChild(msgDiv);
                        return;
                    }}

                    for (let i = 0; i < total; i++) {{
                        let dayN, mo, yr, otro = false;
                        if (i < firstDay) {{
                            dayN = daysInPrev - firstDay + i + 1; mo = viewMonth-1; yr = viewYear; otro = true;
                        }} else if (i >= firstDay + daysInMonth) {{
                            dayN = i - firstDay - daysInMonth + 1; mo = viewMonth+1; yr = viewYear; otro = true;
                        }} else {{
                            dayN = i - firstDay + 1; mo = viewMonth; yr = viewYear;
                        }}
                        if (mo < 0)  {{ mo = 11; yr--; }}
                        if (mo > 11) {{ mo = 0;  yr++; }}

                        const cell = new Date(yr, mo, dayN);
                        cell.setHours(0,0,0,0);
                        const isHoy  = cell.getTime() === todayDate.getTime();
                        const dayEvs = evsByDay(yr, mo, dayN);

                        const div = document.createElement('div');
                        div.className = 'cal-day' + (otro ? ' otro-mes' : '') + (isHoy ? ' hoy' : '');

                        let html = `<div class="day-num">${{dayN}}</div>`;
                        const max = 2;
                        dayEvs.slice(0, max).forEach(ev => {{
                            const cat = ev.categoria;
                            const idx = eventos.indexOf(ev);
                            const nm  = ev.empleado.split(' ')[0];
                            const cur = ev.curso.split(' ').slice(0,2).join(' ');
                            html += `<div class="ev-pill ${{cat}}" title="${{ev.empleado}} – ${{ev.curso}}" onclick="showDetail(${{idx}})">${{nm}}: ${{cur}}</div>`;
                        }});
                        if (dayEvs.length > max) {{
                            html += `<div class="more-tag" onclick="showDayAll('${{yr}}','${{mo}}','${{dayN}}')">+${{dayEvs.length - max}} más</div>`;
                        }}
                        div.innerHTML = html;
                        grid.appendChild(div);
                    }}
                }}

                function renderLista() {{
                    const lv = document.getElementById('listaView');
                    document.getElementById('calPeriod').textContent = 'Lista de eventos';
                    if (!eventos.length) {{
                        lv.innerHTML = '<div style="text-align:center;padding:2rem;color:#aaa;font-size:14px">Sin eventos en el período</div>';
                        return;
                    }}
                    const sorted = [...eventos].sort((a,b) => a.fecha - b.fecha);
                    let html = '';
                    let lastTs = null;
                    sorted.forEach(ev => {{
                        const df = new Date(ev.fecha); df.setHours(0,0,0,0);
                        if (!lastTs || df.getTime() !== lastTs) {{
                            const ds = df.toLocaleDateString('es-CO', {{weekday:'long', day:'numeric', month:'long', year:'numeric'}});
                            html += `<div class="list-date-header">${{ds}}</div>`;
                            lastTs = df.getTime();
                        }}
                        const cat = ev.categoria;
                        html += `<div class="ev-row">
                            <div class="ev-dot" style="background:${{dotColor(cat)}}"></div>
                            <div class="ev-info">
                                <div class="ev-name">${{ev.empleado}}</div>
                                <div class="ev-meta">${{ev.curso}} · ${{ev.cliente}}</div>
                            </div>
                            <span class="ev-badge badge-${{cat}}">${{badgeText(ev)}}</span>
                        </div>`;
                    }});
                    lv.innerHTML = `<div style="background:#fff;border:0.5px solid #ebebeb;border-radius:12px;padding:10px 14px">${{html}}</div>`;
                }}

                function renderAnual() {{
                    document.getElementById('calPeriod').textContent = `${{viewYear}}`;
                    const grid = document.getElementById('anualGrid');
                    grid.innerHTML = '';
                    for (let m = 0; m < 12; m++) {{
                        const mes = document.createElement('div');
                        mes.className = 'mini-mes';
                        const firstDay    = new Date(viewYear, m, 1).getDay();
                        const daysInMonth = new Date(viewYear, m+1, 0).getDate();
                        const daysInPrev  = new Date(viewYear, m, 0).getDate();
                        const total       = Math.ceil((firstDay + daysInMonth) / 7) * 7;
                        let html = `<div class="mini-mes-title">${{MESES[m]}}</div>
                        <div class="mini-week-hdr">${{SEMANAS_CORTO.map(s=>`<span>${{s}}</span>`).join('')}}</div>
                        <div class="mini-grid">`;
                        for (let i = 0; i < total; i++) {{
                            let dayN, mo2, yr2, otro = false;
                            if (i < firstDay) {{ dayN = daysInPrev-firstDay+i+1; mo2 = m-1; yr2 = viewYear; otro = true; }}
                            else if (i >= firstDay+daysInMonth) {{ dayN = i-firstDay-daysInMonth+1; mo2 = m+1; yr2 = viewYear; otro = true; }}
                            else {{ dayN = i-firstDay+1; mo2 = m; yr2 = viewYear; }}
                            if (mo2 < 0)  {{ mo2=11; yr2--; }}
                            if (mo2 > 11) {{ mo2=0;  yr2++; }}
                            const cell = new Date(yr2, mo2, dayN); cell.setHours(0,0,0,0);
                            const isHoy = cell.getTime()===todayDate.getTime();
                            const dayEvs = evsByDay(yr2, mo2, dayN);
                            const hasEv  = dayEvs.length > 0;
                            let cls = 'mini-day';
                            if (otro) cls += ' otro-mes-m';
                            if (isHoy) cls += ' hoy-m';
                            if (hasEv && !otro) cls += ' con-eventos';
                            const onclick = hasEv && !otro ? `onclick="showDayAll('${{yr2}}','${{mo2}}','${{dayN}}')"` : '';
                            const title   = hasEv ? `title="${{dayEvs.length}} evento(s)"` : '';
                            html += `<div class="${{cls}}" ${{onclick}} ${{title}}>${{dayN}}</div>`;
                        }}
                        html += '</div>';
                        mes.innerHTML = html;
                        grid.appendChild(mes);
                    }}
                }}

                function showDetail(idx) {{
                    const ev = eventos[idx];
                    const cat = ev.categoria;
                    document.getElementById('detailTitle').innerHTML = '📌 Detalle del vencimiento';
                    document.getElementById('detailContent').innerHTML = `
                        <div class="detail-content-row">
                            <div class="detail-main">
                                <div class="d-emp">${{ev.empleado}}</div>
                                <div class="d-cur">${{ev.curso}}</div>
                                <div class="d-cli">${{ev.cliente}}</div>
                            </div>
                            <div class="detail-right">
                                <span class="d-badge badge-${{cat}}">${{badgeText(ev)}}</span>
                                <div class="d-date">${{ev.fecha.toLocaleDateString('es-CO',{{day:'numeric',month:'long',year:'numeric'}})}}</div>
                            </div>
                        </div>`;
                    document.getElementById('detailPanel').style.display = 'block';
                }}

                function showDayAll(yr, mo, dayN) {{
                    const dayEvs = evsByDay(parseInt(yr), parseInt(mo), parseInt(dayN));
                    const dateStr = new Date(parseInt(yr),parseInt(mo),parseInt(dayN)).toLocaleDateString('es-CO',{{weekday:'long',day:'numeric',month:'long',year:'numeric'}});
                    document.getElementById('detailTitle').textContent = dateStr.charAt(0).toUpperCase()+dateStr.slice(1);
                    let html = '';
                    dayEvs.forEach(ev => {{
                        const cat = ev.categoria;
                        html += `<div class="ev-row">
                            <div class="ev-dot" style="background:${{dotColor(cat)}}"></div>
                            <div class="ev-info">
                                <div class="ev-name">${{ev.empleado}}</div>
                                <div class="ev-meta">${{ev.curso}} · ${{ev.cliente}}</div>
                            </div>
                            <span class="ev-badge badge-${{cat}}">${{badgeText(ev)}}</span>
                        </div>`;
                    }});
                    document.getElementById('detailContent').innerHTML = html;
                    document.getElementById('detailPanel').style.display = 'block';
                }}

                function setView(v) {{
                    currentView = v;
                    ['btnMes','btnLista','btnAnual'].forEach(id => document.getElementById(id).classList.remove('active'));
                    document.getElementById('btn'+v.charAt(0).toUpperCase()+v.slice(1)).classList.add('active');
                    document.getElementById('mesView').style.display   = v==='mes'   ? 'block' : 'none';
                    document.getElementById('listaView').style.display = v==='lista' ? 'block' : 'none';
                    document.getElementById('anualView').style.display = v==='anual' ? 'block' : 'none';
                    document.getElementById('detailPanel').style.display = 'none';
                    if (v==='mes')   renderMes();
                    if (v==='lista') renderLista();
                    if (v==='anual') renderAnual();
                }}

                renderStats();
                renderMes();
                </script>
                </body>
                </html>
                """

                # Altura dinámica
                cal_height = 820 if tipo_filtro in ("Año",) else 680

                # Mostrar el iframe
                st.iframe(calendario_html, height=cal_height)

                # 🟢 TABLA DE DETALLE 
                st.markdown("---")
                st.subheader("📋 Detalle de cursos próximos a vencer")
                
                # ---- Filtro por empleado ----
                empleados_opciones = ["Todos"] + sorted(por_vencer['empleado'].unique().tolist())
                empleado_filtro = st.selectbox("Filtrar por empleado", empleados_opciones, key="filtro_empleado_calendario")
                
                df_detalle = por_vencer.copy()
                if empleado_filtro != "Todos":
                    df_detalle = df_detalle[df_detalle['empleado'] == empleado_filtro]
                
                df_detalle = df_detalle.sort_values('fecha_vencimiento')
                
                if df_detalle.empty:
                    st.info("No hay eventos para este filtro.")
                else:
                    df_detalle['fecha_vencimiento_mostrar'] = df_detalle['fecha_vencimiento'].dt.strftime("%d/%m/%Y")
                    df_detalle['dias_restantes'] = (df_detalle['fecha_vencimiento'] - pd.Timestamp(date.today())).dt.days
                    df_detalle['categoria'] = df_detalle.apply(categoria_calendario, axis=1)

                    ETIQUETAS_CATEGORIA = {
                        'vencido': '🔴 Vencido (incumplimiento)',
                        'urgente': '🔴 Urgente (≤3 días)',
                        'proximo': '🟡 Próximo (≤30 días)',
                        'normal': '🔵 Vigente (>30 días)',
                        'sin_vencimiento': '⚪ Sin vencimiento',
                    }
                    df_detalle['Categoría'] = df_detalle['categoria'].map(ETIQUETAS_CATEGORIA)
                    
                    def color_dias(dias):
                        if dias < 0:
                            return "background-color: #F8D7DA; color: #721C24; font-weight: bold;"
                        elif dias <= 3:
                            return "background-color: #F8D7DA; color: #721C24;"
                        elif dias <= 30:
                            return "background-color: #FFF3CD; color: #856404;"
                        else:
                            return "background-color: #D1ECF1; color: #0C5460;"
                    
                    df_detalle['color'] = df_detalle['dias_restantes'].apply(color_dias)
                    
                    st.dataframe(
                        df_detalle[['empleado', 'cedula', 'curso', 'cliente', 'Categoría', 'fecha_vencimiento_mostrar', 'dias_restantes']].style.apply(
                            lambda x: [df_detalle.loc[x.name, 'color'] for _ in x], axis=1
                        ),
                        column_config={
                            "empleado": "Empleado",
                            "cedula": "Cédula",
                            "curso": "Curso",
                            "cliente": "Cliente",
                            "Categoría": "Categoría",
                            "fecha_vencimiento_mostrar": "Fecha vencimiento",
                            "dias_restantes": st.column_config.NumberColumn("Días restantes", format="%d")
                        },
                        hide_index=True,
                        width='stretch'
                    )
                    
                    # ---- Descarga en Excel (corregido) ----
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                    def generar_excel_alertas(df_detalle):
                        # Resetear índice para evitar filas en blanco
                        df = df_detalle.reset_index(drop=True)
                        
                        output = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Alertas de Vencimiento"
                        
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                        center_align = Alignment(horizontal="center", vertical="center")
                        
                        # Encabezados
                        headers = ['Empleado', 'Cédula', 'Curso', 'Cliente', 'Fecha vencimiento', 'Días restantes']
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = border
                        
                        # Datos - usando enumerate con start=2 para filas consecutivas
                        for i, row in df.iterrows():
                            row_num = i + 2  # i comienza en 0, así la primera fila de datos va en la 2
                            ws.cell(row=row_num, column=1, value=row['empleado']).border = border
                            ws.cell(row=row_num, column=2, value=row['cedula']).border = border
                            ws.cell(row=row_num, column=3, value=row['curso']).border = border
                            ws.cell(row=row_num, column=4, value=row['cliente']).border = border
                            ws.cell(row=row_num, column=5, value=row['fecha_vencimiento_mostrar']).border = border
                            
                            dias = row['dias_restantes']
                            cell_dias = ws.cell(row=row_num, column=6, value=dias)
                            cell_dias.border = border
                            cell_dias.alignment = center_align
                            # Color según urgencia
                            if dias <= 7:
                                cell_dias.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                            elif dias <= 15:
                                cell_dias.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            else:
                                cell_dias.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                        
                        # Ajustar anchos de columna
                        ws.column_dimensions['A'].width = 30
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 30
                        ws.column_dimensions['D'].width = 20
                        ws.column_dimensions['E'].width = 18
                        ws.column_dimensions['F'].width = 15
                        
                        wb.save(output)
                        output.seek(0)
                        return output

                    # Generar Excel con los datos filtrados
                    excel_data = generar_excel_alertas(df_detalle)

                    st.download_button(
                        label="📥 Descargar alertas en Excel",
                        data=excel_data,
                        file_name=f'alertas_vencimiento_{date.today().strftime("%Y%m%d")}.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key='download_alerts_excel'
                    )

                    st.markdown("---")


    with tabs[6]:
            st.subheader("📈 Cumplimiento por Periodo (Evolución Mensual)")

            from dateutil.relativedelta import relativedelta

            asignaciones = db.obtener_asignaciones()
            if asignaciones.empty:
                st.info("No hay asignaciones para mostrar.")
            else:
                asignaciones['fecha_asignacion']  = pd.to_datetime(asignaciones['fecha_asignacion'])
                asignaciones['fecha_vencimiento'] = pd.to_datetime(asignaciones['fecha_vencimiento'])

                # ── Utilidades DAX-equivalentes ──────────────────────────────────────
                ESTADOS_EXCLUIR  = {"NO APLICA", "RESTRICCIÓN", "sin_ejecucion"}

                def calcular_tasa_mes(df, fecha_ref):
                    ultimo_dia = fecha_ref + pd.offsets.MonthEnd(0)
                    activas = df[
                        (df['fecha_asignacion'] <= ultimo_dia) &
                        (~df['estado'].isin(ESTADOS_EXCLUIR))
                    ]
                    en_regla = activas[
                        (activas['estado'] == 'completado') |
                        (
                            (activas['estado'] == 'pendiente') &
                            (activas['fecha_vencimiento'] >= ultimo_dia)
                        )
                    ]
                    total = len(activas)
                    return round(len(en_regla) / total * 100, 1) if total > 0 else 0.0

                def indicador_variacion(tasa_actual, tasa_anterior):
                    dif = tasa_actual - tasa_anterior
                    if   dif >=  0.1:  simbolo = "▲ +"
                    elif dif <= -0.1:  simbolo = "▼ "
                    else:               simbolo = "▬ "
                    return simbolo + f"{abs(dif):.1f}%", dif

                # ── Período derivado de los propios datos ────────────────────────────
                ultima_fecha   = asignaciones['fecha_asignacion'].max()
                mes_actual_ref = ultima_fecha.replace(day=1)
                mes_ant_ref    = mes_actual_ref - relativedelta(months=1)

                tasa_ultimo   = calcular_tasa_mes(asignaciones, mes_actual_ref)
                tasa_anterior = calcular_tasa_mes(asignaciones, mes_ant_ref)
                texto_var, dif_var = indicador_variacion(tasa_ultimo, tasa_anterior)

                # ── Cabecera: métricas de período ────────────────────────────────────
                st.markdown("##### Período derivado de los datos")
                h1, h2, h3, h4 = st.columns(4)
                h1.metric(
                    "📅 Último período con datos",
                    ultima_fecha.strftime("%B %Y").capitalize()
                )
                h2.metric(
                    "✅ Cumplimiento último mes",
                    f"{tasa_ultimo:.1f}%",
                    delta=texto_var,
                    delta_color="normal" if dif_var >= 0 else "inverse"
                )
                h3.metric(
                    "📊 Mes anterior",
                    f"{tasa_anterior:.1f}%",
                    help=f"Período: {mes_ant_ref.strftime('%B %Y').capitalize()}"
                )
                h4.metric(
                    "📉 Variación",
                    texto_var,
                    help="▲ mejora · ▼ baja · ▬ estable"
                )

                st.markdown("---")

                # ── Filtros ──────────────────────────────────────────────────────────
                col_f1, col_f2, col_f3 = st.columns(3)
                with col_f1:
                    clientes = sorted(asignaciones['cliente'].unique())
                    cliente_sel = st.selectbox("Filtrar por cliente", ["Todos"] + clientes, key="periodo_cliente_dax")
                with col_f2:
                    if cliente_sel != "Todos":
                        cursos_disponibles = sorted(asignaciones[asignaciones['cliente'] == cliente_sel]['curso'].unique())
                    else:
                        cursos_disponibles = sorted(asignaciones['curso'].unique())
                    curso_sel = st.selectbox("Filtrar por curso (opcional)", ["Todos"] + cursos_disponibles, key="periodo_curso_dax")
                with col_f3:
                    opciones_periodo = {
                        "Último mes":         1,
                        "Últimos 2 meses":    2,
                        "Últimos 3 meses":    3,
                        "Últimos 6 meses":    6,
                        "Último año":        12,
                        "Últimos 2 años":    24,
                        "Últimos 3 años":    36,
                        "Últimos 5 años":    60,
                        "Todo el histórico": -1
                    }
                    periodo_sel    = st.selectbox("Periodo de análisis", list(opciones_periodo.keys()), index=0, key="periodo_meses_dax")
                    meses_analisis = opciones_periodo[periodo_sel]

                # ── Aplicar filtros ──────────────────────────────────────────────────
                df_filtrado = asignaciones.copy()
                if cliente_sel != "Todos":
                    df_filtrado = df_filtrado[df_filtrado['cliente'] == cliente_sel]
                if curso_sel != "Todos":
                    df_filtrado = df_filtrado[df_filtrado['curso'] == curso_sel]

                if df_filtrado.empty:
                    st.warning("No hay datos con los filtros seleccionados.")
                else:
                    hoy = pd.Timestamp(date.today())

                    # ── Rango de meses desde los datos del filtro activo ─────────────
                    ultima_fecha_filtrado  = df_filtrado['fecha_asignacion'].max()
                    primera_fecha_filtrado = df_filtrado['fecha_asignacion'].min()

                    if meses_analisis == -1:
                        min_date = primera_fecha_filtrado.replace(day=1)
                        max_date = ultima_fecha_filtrado.replace(day=1)
                    else:
                        max_date = ultima_fecha_filtrado.replace(day=1)
                        min_date = (max_date - pd.DateOffset(months=meses_analisis - 1)).replace(day=1)
                        if primera_fecha_filtrado.replace(day=1) > min_date:
                            meses_disponibles = len(pd.date_range(
                                start=primera_fecha_filtrado.replace(day=1),
                                end=max_date, freq='MS'
                            ))
                            st.info(f"Solo hay {meses_disponibles} meses disponibles para este filtro.")
                            min_date = primera_fecha_filtrado.replace(day=1)

                    meses = pd.date_range(start=min_date, end=max_date, freq='MS')

                    # ── Serie mensual con lógica DAX ─────────────────────────────────
                    resultados = []
                    for mes in meses:
                        tasa = calcular_tasa_mes(df_filtrado, mes)
                        resultados.append({'mes': mes, 'tasa': tasa})

                    df_meses = pd.DataFrame(resultados)
                    df_meses['variacion'] = df_meses['tasa'].diff().round(1)
                    df_meses['variacion_formateada'] = df_meses.apply(
                        lambda r: (
                            f"▲ +{r['variacion']:.1f}%" if r['variacion'] >= 0.1
                            else (f"▼ {r['variacion']:.1f}%" if r['variacion'] <= -0.1
                                else "▬ 0.0%")
                        ) if pd.notna(r['variacion']) else "",
                        axis=1
                    )

                    # ── KPIs resumen ─────────────────────────────────────────────────
                    ultimo   = df_meses.iloc[-1]
                    primero  = df_meses.iloc[0]
                    var_tot  = ultimo['tasa'] - primero['tasa']
                    promedio = df_meses['tasa'].mean()

                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("📊 Tasa actual",       f"{ultimo['tasa']:.1f}%")
                    k2.metric("📈 Variación total",    f"{var_tot:+.1f}%",
                            delta_color="normal" if var_tot >= 0 else "inverse")
                    k3.metric("📅 Meses analizados",   len(df_meses))
                    k4.metric("🎯 Promedio histórico", f"{promedio:.1f}%")

                    # ── Gráfica ──────────────────────────────────────────────────────
                    fig = px.line(
                        df_meses, x='mes', y='tasa',
                        title=f'Evolución de la tasa de cumplimiento ({periodo_sel})',
                        labels={'mes': 'Mes', 'tasa': 'Tasa (%)'},
                        markers=True,
                        custom_data=['variacion_formateada']
                    )
                    fig.update_traces(
                        hovertemplate="<b>%{x|%B %Y}</b><br>Tasa: %{y:.1f}%<br>Variación: %{customdata[0]}<extra></extra>"
                    )
                    fig.add_hline(
                        y=promedio, line_dash="dash", line_color="blue",
                        annotation_text=f"Promedio: {promedio:.1f}%",
                        annotation_font_color="blue"
                    )
                    fig.update_layout(height=450, yaxis_range=[0, 105])
                    st.plotly_chart(fig, width='stretch')

                    # ── Tabla mensual ────────────────────────────────────────────────
                    st.subheader("📋 Detalle mensual")
                    df_tabla = df_meses[['mes', 'tasa', 'variacion_formateada']].copy()
                    df_tabla['mes'] = df_tabla['mes'].dt.strftime('%B %Y')
                    df_tabla = df_tabla.rename(columns={
                        'mes':                  'Mes',
                        'tasa':                 'Tasa %',
                        'variacion_formateada': 'Variación'
                    })

                    def color_variacion(val):
                        if isinstance(val, str):
                            if val.startswith('▲'): return 'color: #28a745; font-weight: bold;'
                            if val.startswith('▼'): return 'color: #dc3545; font-weight: bold;'
                        return ''

                    st.dataframe(
                        df_tabla.style.map(color_variacion, subset=['Variación']),
                        hide_index=True,
                        width='stretch'
                    )

                    # ── Análisis automático ──────────────────────────────────────────
                    if len(df_meses) > 2:
                        mejor_mes = df_meses.loc[df_meses['tasa'].idxmax()]
                        peor_mes  = df_meses.loc[df_meses['tasa'].idxmin()]
                        tendencia = (
                            '⬆️ Creciente' if var_tot > 0
                            else ('⬇️ Decreciente' if var_tot < 0 else '➡️ Estable')
                        )

                        st.markdown(f"""
                        **📌 Análisis ({periodo_sel}):**  
                        - Mejor mes: **{mejor_mes['mes'].strftime('%B %Y')}** — {mejor_mes['tasa']:.1f}%  
                        - Peor mes:  **{peor_mes['mes'].strftime('%B %Y')}** — {peor_mes['tasa']:.1f}%  
                        - Tendencia: {tendencia} ({var_tot:+.1f}% desde el inicio del periodo)  
                        - Promedio histórico: **{promedio:.1f}%**
                        """)

                        if ultimo['tasa'] < promedio:
                            st.warning("⚠️ La tasa actual está por debajo del promedio histórico.")
                        elif ultimo['tasa'] > promedio * 1.1:
                            st.success("✅ La tasa actual supera el promedio histórico en más del 10%.")
                    else:
                        st.info("Se necesitan al menos 3 meses de datos para el análisis de tendencia.") 